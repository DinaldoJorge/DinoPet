# ==========================================================
# DINOPETANIMAL PRO 5.0
# ARQUIVO:
# excel.py
#
# PARTE 01/20
# IMPORTS E CONFIGURAÇÃO
# ==========================================================

from pathlib import Path

from datetime import datetime

from shutil import copy2

from openpyxl import load_workbook

# ==========================================================
# DIRETÓRIO BASE
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent

# ==========================================================
# PASTAS
# ==========================================================

PASTA_PETS = BASE_DIR / "pets"

PASTA_RELATORIOS = BASE_DIR / "relatorios"

# ==========================================================
# ARQUIVOS
# ==========================================================

MODELO_PET = BASE_DIR / "modelo_pet.xlsx"

# ==========================================================
# CRIA PASTAS NECESSÁRIAS
# ==========================================================

PASTA_PETS.mkdir(

    parents=True,

    exist_ok=True

)

PASTA_RELATORIOS.mkdir(

    parents=True,

    exist_ok=True

)

# ==========================================================
# VERIFICA O MODELO
# ==========================================================

if not MODELO_PET.exists():

    raise FileNotFoundError(

        f"Modelo não encontrado:\n{MODELO_PET}"

    )

# ==========================================================
# DATA E HORA
# ==========================================================

def agora():

    """
    Retorna data e hora atual.
    """

    return datetime.now()

# ==========================================================
# DATA FORMATADA
# ==========================================================

def agora_texto():

    """
    Retorna a data e hora formatadas.
    """

    return agora().strftime(

        "%d/%m/%Y %H:%M:%S"

    )

# ==========================================================
# CAMINHO DO PET
# ==========================================================

def caminho_pet(nome):

    """
    Retorna o caminho do arquivo Excel do pet.
    """

    nome = str(nome).strip()

    return PASTA_PETS / f"{nome}.xlsx"

# ==========================================================
# FIM DA PARTE 01/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 02/20
# FUNÇÕES AUXILIARES
# ==========================================================

# ==========================================================
# VERIFICA SE O PET EXISTE
# ==========================================================

def pet_existe(nome):

    """
    Verifica se existe um arquivo Excel
    para o pet informado.
    """

    return caminho_pet(nome).exists()

# ==========================================================
# LISTA TODOS OS PETS
# ==========================================================

def listar_pets():

    """
    Retorna uma lista com todos
    os pets cadastrados.
    """

    lista = []

    for arquivo in sorted(

        PASTA_PETS.glob("*.xlsx")

    ):

        lista.append(

            arquivo.stem

        )

    return lista

# ==========================================================
# TOTAL DE PETS
# ==========================================================

def total_pets():

    """
    Quantidade total de pets.
    """

    return len(

        listar_pets()

    )

# ==========================================================
# ÚLTIMO PET
# ==========================================================

def ultimo_pet():

    """
    Retorna o último pet cadastrado.
    """

    lista = listar_pets()

    if not lista:

        return "-"

    return lista[-1]

# ==========================================================
# ABRE PLANILHA
# ==========================================================

def abrir_planilha(nome):

    """
    Abre o arquivo Excel do pet.
    """

    arquivo = caminho_pet(nome)

    if not arquivo.exists():

        raise FileNotFoundError(

            f"Pet não encontrado: {nome}"

        )

    return load_workbook(

        arquivo

    )

# ==========================================================
# SALVA PLANILHA
# ==========================================================

def salvar_planilha(

    workbook,

    nome

):

    """
    Salva a planilha.
    """

    workbook.save(

        caminho_pet(nome)

    )

# ==========================================================
# COPIA MODELO
# ==========================================================

def copiar_modelo(nome):

    """
    Copia modelo_pet.xlsx
    para um novo paciente.
    """

    destino = caminho_pet(nome)

    copy2(

        MODELO_PET,

        destino

    )

    return destino

# ==========================================================
# REMOVE ESPAÇOS
# ==========================================================

def limpar_texto(texto):

    """
    Remove espaços extras.
    """

    return str(

        texto

    ).strip()

# ==========================================================
# NOME VÁLIDO
# ==========================================================

def nome_valido(nome):

    """
    Validação simples.
    """

    return len(

        limpar_texto(nome)

    ) >= 2

# ==========================================================
# FIM DA PARTE 02/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 03/20
# CRIAR NOVO PET
# ==========================================================

# ==========================================================
# CRIAR PET
# ==========================================================

def criar_pet(dados):

    """
    Cria um novo paciente utilizando
    o modelo_pet.xlsx.
    """

    nome = limpar_texto(

        dados.get(

            "nome",

            ""

        )

    )

    # ------------------------------------------------------

    if not nome_valido(nome):

        raise ValueError(

            "Nome inválido."

        )

    # ------------------------------------------------------

    if pet_existe(nome):

        raise FileExistsError(

            f"O pet '{nome}' já existe."

        )

    # ------------------------------------------------------

    copiar_modelo(nome)

    wb = abrir_planilha(nome)

    ws = wb["Cadastro"]

    # ======================================================
    # DADOS BÁSICOS
    # ======================================================

    ws["B2"] = agora_texto()

    ws["B4"] = nome

    ws["B5"] = dados.get(

        "especie",

        ""

    )

    ws["B6"] = dados.get(

        "raca",

        ""

    )

    ws["B7"] = dados.get(

        "sexo",

        ""

    )

    ws["B8"] = dados.get(

        "idade",

        ""

    )

    ws["B9"] = dados.get(

        "nascimento",

        ""

    )

    # ======================================================
    # TUTOR
    # ======================================================

    ws["B11"] = dados.get(

        "tutor",

        ""

    )

    ws["B12"] = dados.get(

        "telefone",

        ""

    )

    ws["B13"] = dados.get(

        "email",

        ""

    )

    # ======================================================
    # DADOS FÍSICOS
    # ======================================================

    ws["B15"] = dados.get(

        "peso",

        ""

    )

    ws["B16"] = dados.get(

        "altura",

        ""

    )

    ws["B17"] = dados.get(

        "temperatura",

        ""

    )

    # ======================================================
    # VETERINÁRIO
    # ======================================================

    ws["B19"] = dados.get(

        "veterinario",

        ""

    )

    ws["B20"] = dados.get(

        "clinica",

        ""

    )

    # ======================================================
    # OBSERVAÇÕES
    # ======================================================

    ws["B22"] = dados.get(

        "observacoes",

        ""

    )

    salvar_planilha(

        wb,

        nome

    )

    return True

# ==========================================================
# FIM DA PARTE 03/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 04/20
# ABRIR PET E LER CADASTRO
# ==========================================================

# ==========================================================
# LER CADASTRO
# ==========================================================

def ler_cadastro(nome):

    """
    Lê a aba Cadastro e devolve
    um dicionário com os dados do pet.
    """

    if not pet_existe(nome):

        raise FileNotFoundError(

            f"Pet '{nome}' não encontrado."

        )

    wb = abrir_planilha(nome)

    ws = wb["Cadastro"]

    dados = {

        "nome": ws["B4"].value,
        "especie": ws["B5"].value,
        "raca": ws["B6"].value,
        "sexo": ws["B7"].value,
        "idade": ws["B8"].value,
        "nascimento": ws["B9"].value,

        "tutor": ws["B11"].value,
        "telefone": ws["B12"].value,
        "email": ws["B13"].value,

        "peso": ws["B15"].value,
        "altura": ws["B16"].value,
        "temperatura": ws["B17"].value,

        "veterinario": ws["B19"].value,
        "clinica": ws["B20"].value,

        "observacoes": ws["B22"].value

    }

    wb.close()

    return dados

# ==========================================================
# ABRIR PET
# ==========================================================

def abrir_pet(nome):

    """
    Função utilizada pelas páginas
    do sistema.
    """

    return ler_cadastro(nome)

# ==========================================================
# RESUMO DO PET
# ==========================================================

def resumo_pet(nome):

    """
    Retorna um resumo simples
    do paciente.
    """

    dados = ler_cadastro(nome)

    return {

        "Nome": dados.get("nome"),

        "Espécie": dados.get("especie"),

        "Raça": dados.get("raca"),

        "Tutor": dados.get("tutor"),

        "Peso": dados.get("peso")

    }

# ==========================================================
# FIM DA PARTE 04/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 05/20
# ATUALIZAR CADASTRO
# ==========================================================

# ==========================================================
# ATUALIZAR PET
# ==========================================================

def atualizar_pet(dados):

    """
    Atualiza os dados de um pet existente.
    """

    nome = limpar_texto(

        dados.get(

            "nome",

            ""

        )

    )

    if not pet_existe(nome):

        raise FileNotFoundError(

            f"Pet '{nome}' não encontrado."

        )

    wb = abrir_planilha(nome)

    ws = wb["Cadastro"]

    # ======================================================
    # DATA DE ALTERAÇÃO
    # ======================================================

    ws["B3"] = agora_texto()

    # ======================================================
    # DADOS BÁSICOS
    # ======================================================

    ws["B4"] = dados.get("nome", "")
    ws["B5"] = dados.get("especie", "")
    ws["B6"] = dados.get("raca", "")
    ws["B7"] = dados.get("sexo", "")
    ws["B8"] = dados.get("idade", "")
    ws["B9"] = dados.get("nascimento", "")

    # ======================================================
    # TUTOR
    # ======================================================

    ws["B11"] = dados.get("tutor", "")
    ws["B12"] = dados.get("telefone", "")
    ws["B13"] = dados.get("email", "")

    # ======================================================
    # DADOS FÍSICOS
    # ======================================================

    ws["B15"] = dados.get("peso", "")
    ws["B16"] = dados.get("altura", "")
    ws["B17"] = dados.get("temperatura", "")

    # ======================================================
    # VETERINÁRIO
    # ======================================================

    ws["B19"] = dados.get("veterinario", "")
    ws["B20"] = dados.get("clinica", "")

    # ======================================================
    # OBSERVAÇÕES
    # ======================================================

    ws["B22"] = dados.get("observacoes", "")

    salvar_planilha(

        wb,

        nome

    )

    wb.close()

    return True

# ==========================================================
# ALTERAR NOME DO PET
# ==========================================================

def renomear_pet(

    nome_antigo,

    nome_novo

):

    """
    Renomeia o arquivo Excel do pet.
    """

    nome_antigo = limpar_texto(

        nome_antigo

    )

    nome_novo = limpar_texto(

        nome_novo

    )

    origem = caminho_pet(

        nome_antigo

    )

    destino = caminho_pet(

        nome_novo

    )

    if not origem.exists():

        raise FileNotFoundError(

            nome_antigo

        )

    if destino.exists():

        raise FileExistsError(

            nome_novo

        )

    origem.rename(

        destino

    )

    return True

# ==========================================================
# FIM DA PARTE 05/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 06/20
# EXCLUIR PET
# ==========================================================

# ==========================================================
# EXCLUIR PET
# ==========================================================

def excluir_pet(nome):

    """
    Exclui definitivamente o arquivo
    Excel do pet.
    """

    nome = limpar_texto(nome)

    arquivo = caminho_pet(nome)

    if not arquivo.exists():

        raise FileNotFoundError(

            f"Pet '{nome}' não encontrado."

        )

    arquivo.unlink()

    return True

# ==========================================================
# EXCLUIR RELATÓRIO
# ==========================================================

def excluir_relatorio(nome):

    """
    Remove o relatório PDF do pet,
    caso exista.
    """

    nome = limpar_texto(nome)

    pdf = PASTA_RELATORIOS / f"{nome}.pdf"

    if pdf.exists():

        pdf.unlink()

        return True

    return False

# ==========================================================
# EXCLUIR PET COMPLETO
# ==========================================================

def excluir_pet_completo(nome):

    """
    Remove o cadastro e também
    o relatório PDF.
    """

    excluir_pet(nome)

    excluir_relatorio(nome)

    return True

# ==========================================================
# LIMPAR TODOS OS RELATÓRIOS
# ==========================================================

def limpar_relatorios():

    """
    Remove todos os PDFs.
    """

    quantidade = 0

    for arquivo in PASTA_RELATORIOS.glob("*.pdf"):

        arquivo.unlink()

        quantidade += 1

    return quantidade

# ==========================================================
# LIMPAR TODOS OS PETS
# ==========================================================

def limpar_pets():

    """
    Remove todos os pacientes.
    """

    quantidade = 0

    for arquivo in PASTA_PETS.glob("*.xlsx"):

        arquivo.unlink()

        quantidade += 1

    return quantidade

# ==========================================================
# FIM DA PARTE 06/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 07/20
# LISTAGEM E PESQUISA
# ==========================================================

# ==========================================================
# PESQUISAR PETS
# ==========================================================

def pesquisar_pets(texto=""):

    """
    Pesquisa pets pelo nome.
    """

    texto = limpar_texto(

        texto

    ).lower()

    resultado = []

    for nome in listar_pets():

        if texto in nome.lower():

            resultado.append(

                nome

            )

    return sorted(

        resultado

    )

# ==========================================================
# LISTAR DADOS DOS PETS
# ==========================================================

def listar_dados_pets():

    """
    Retorna uma lista contendo
    os dados básicos de todos os pets.
    """

    lista = []

    for nome in listar_pets():

        try:

            dados = ler_cadastro(

                nome

            )

            lista.append(

                dados

            )

        except Exception:

            pass

    return lista

# ==========================================================
# PET MAIS RECENTE
# ==========================================================

def pet_mais_recente():

    """
    Retorna o pet cadastrado
    mais recentemente.
    """

    arquivos = sorted(

        PASTA_PETS.glob("*.xlsx"),

        key=lambda arq: arq.stat().st_mtime,

        reverse=True

    )

    if not arquivos:

        return None

    return arquivos[0].stem

# ==========================================================
# QUANTIDADE POR ESPÉCIE
# ==========================================================

def total_por_especie():

    """
    Conta quantos pets existem
    por espécie.
    """

    resultado = {}

    for pet in listar_dados_pets():

        especie = pet.get(

            "especie",

            "Não informado"

        )

        resultado[especie] = resultado.get(

            especie,

            0

        ) + 1

    return resultado

# ==========================================================
# QUANTIDADE POR SEXO
# ==========================================================

def total_por_sexo():

    """
    Conta machos e fêmeas.
    """

    resultado = {

        "Macho": 0,

        "Fêmea": 0

    }

    for pet in listar_dados_pets():

        sexo = pet.get(

            "sexo",

            ""

        )

        if sexo in resultado:

            resultado[sexo] += 1

    return resultado

# ==========================================================
# FIM DA PARTE 07/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 08/20
# LEITURA E ESCRITA
# ==========================================================

# ==========================================================
# LER CÉLULA
# ==========================================================

def ler(ws, celula, padrao=""):

    """
    Lê uma célula da planilha.
    """

    valor = ws[celula].value

    if valor is None:

        return padrao

    return valor

# ==========================================================
# ESCREVER CÉLULA
# ==========================================================

def escrever(

    ws,

    celula,

    valor

):

    """
    Escreve uma célula.
    """

    ws[celula] = valor

# ==========================================================
# LER VÁRIAS CÉLULAS
# ==========================================================

def ler_campos(

    ws,

    mapa

):

    """
    Lê várias células de uma vez.

    mapa = {

        "nome":"B4",

        "especie":"B5"

    }

    """

    dados = {}

    for campo, celula in mapa.items():

        dados[campo] = ler(

            ws,

            celula

        )

    return dados

# ==========================================================
# ESCREVER VÁRIAS CÉLULAS
# ==========================================================

def escrever_campos(

    ws,

    mapa,

    dados

):

    """
    Escreve vários campos.

    """

    for campo, celula in mapa.items():

        escrever(

            ws,

            celula,

            dados.get(

                campo,

                ""

            )

        )

# ==========================================================
# MAPA DO CADASTRO
# ==========================================================

MAPA_CADASTRO = {

    "nome":"B4",

    "especie":"B5",

    "raca":"B6",

    "sexo":"B7",

    "idade":"B8",

    "nascimento":"B9",

    "tutor":"B11",

    "telefone":"B12",

    "email":"B13",

    "peso":"B15",

    "altura":"B16",

    "temperatura":"B17",

    "veterinario":"B19",

    "clinica":"B20",

    "observacoes":"B22"

}

# ==========================================================
# LER CADASTRO RÁPIDO
# ==========================================================

def ler_cadastro_rapido(nome):

    """
    Lê o cadastro utilizando
    o mapa de células.
    """

    wb = abrir_planilha(

        nome

    )

    ws = wb["Cadastro"]

    dados = ler_campos(

        ws,

        MAPA_CADASTRO

    )

    wb.close()

    return dados

# ==========================================================
# FIM DA PARTE 08/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 09/20
# SALVAR CADASTRO
# ==========================================================

# ==========================================================
# SALVAR CADASTRO
# ==========================================================

def salvar_cadastro(nome, dados):

    """
    Salva todo o cadastro do pet.
    """

    nome = limpar_texto(nome)

    if not pet_existe(nome):

        raise FileNotFoundError(

            f"Pet '{nome}' não encontrado."

        )

    wb = abrir_planilha(nome)

    ws = wb["Cadastro"]

    # ======================================================
    # DATA DA ALTERAÇÃO
    # ======================================================

    escrever(

        ws,

        "B3",

        agora_texto()

    )

    # ======================================================
    # DADOS
    # ======================================================

    escrever_campos(

        ws,

        MAPA_CADASTRO,

        dados

    )

    salvar_planilha(

        wb,

        nome

    )

    wb.close()

    return True

# ==========================================================
# CARREGAR CADASTRO
# ==========================================================

def carregar_cadastro(nome):

    """
    Carrega todo o cadastro.
    """

    return ler_cadastro_rapido(

        nome

    )

# ==========================================================
# DUPLICAR PET
# ==========================================================

def duplicar_pet(

    origem,

    destino

):

    """
    Duplica um paciente.
    """

    origem = limpar_texto(

        origem

    )

    destino = limpar_texto(

        destino

    )

    if not pet_existe(origem):

        raise FileNotFoundError(

            origem

        )

    if pet_existe(destino):

        raise FileExistsError(

            destino

        )

    copy2(

        caminho_pet(origem),

        caminho_pet(destino)

    )

    return True

# ==========================================================
# DATA DA ÚLTIMA ALTERAÇÃO
# ==========================================================

def ultima_alteracao(nome):

    """
    Retorna a data de alteração.
    """

    wb = abrir_planilha(

        nome

    )

    ws = wb["Cadastro"]

    data = ler(

        ws,

        "B3"

    )

    wb.close()

    return data

# ==========================================================
# FIM DA PARTE 09/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 10/20
# AVALIAÇÕES
# ==========================================================

# ==========================================================
# SALVAR AVALIAÇÃO
# ==========================================================

def salvar_avaliacao(

    nome,

    aba,

    dados

):

    """
    Salva uma avaliação em qualquer aba
    da planilha.
    """

    if not pet_existe(nome):

        raise FileNotFoundError(

            nome

        )

    wb = abrir_planilha(nome)

    if aba not in wb.sheetnames:

        raise ValueError(

            f"Aba '{aba}' não encontrada."

        )

    ws = wb[aba]

    for celula, valor in dados.items():

        escrever(

            ws,

            celula,

            valor

        )

    salvar_planilha(

        wb,

        nome

    )

    wb.close()

    return True

# ==========================================================
# LER AVALIAÇÃO
# ==========================================================

def ler_avaliacao(

    nome,

    aba,

    mapa

):

    """
    Lê uma avaliação utilizando
    um mapa de células.
    """

    wb = abrir_planilha(nome)

    if aba not in wb.sheetnames:

        wb.close()

        return {}

    ws = wb[aba]

    dados = ler_campos(

        ws,

        mapa

    )

    wb.close()

    return dados

# ==========================================================
# EXISTE AVALIAÇÃO
# ==========================================================

def existe_avaliacao(

    nome,

    aba

):

    """
    Verifica se a aba existe.
    """

    wb = abrir_planilha(

        nome

    )

    existe = aba in wb.sheetnames

    wb.close()

    return existe

# ==========================================================
# LISTAR AVALIAÇÕES
# ==========================================================

def listar_avaliacoes(nome):

    """
    Lista todas as abas
    da planilha.
    """

    wb = abrir_planilha(

        nome

    )

    abas = wb.sheetnames

    wb.close()

    return abas

# ==========================================================
# FIM DA PARTE 10/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 11/20
# PRONTUÁRIO VETERINÁRIO
# ==========================================================

# ==========================================================
# SALVAR CONSULTA
# ==========================================================

def salvar_consulta(

    nome,

    consulta

):

    """
    Salva os dados da consulta
    na aba PRONTUARIO.
    """

    return salvar_avaliacao(

        nome,

        "PRONTUARIO",

        consulta

    )

# ==========================================================
# LER CONSULTA
# ==========================================================

def ler_consulta(nome):

    """
    Lê o prontuário do pet.
    """

    mapa = {

        "data":"B2",

        "veterinario":"B3",

        "queixa":"B4",

        "diagnostico":"B5",

        "tratamento":"B6",

        "medicamentos":"B7",

        "retorno":"B8",

        "observacoes":"B9"

    }

    return ler_avaliacao(

        nome,

        "PRONTUARIO",

        mapa

    )

# ==========================================================
# ÚLTIMA CONSULTA
# ==========================================================

def ultima_consulta(nome):

    """
    Retorna a data da última consulta.
    """

    dados = ler_consulta(

        nome

    )

    return dados.get(

        "data",

        ""

    )

# ==========================================================
# EXISTE PRONTUÁRIO
# ==========================================================

def existe_prontuario(nome):

    """
    Verifica se existe
    a aba PRONTUARIO.
    """

    return existe_avaliacao(

        nome,

        "PRONTUARIO"

    )

# ==========================================================
# FIM DA PARTE 11/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 12/20
# VACINAÇÃO
# ==========================================================

# ==========================================================
# SALVAR VACINAÇÃO
# ==========================================================

def salvar_vacinacao(

    nome,

    dados

):

    """
    Salva os dados da vacinação.
    """

    return salvar_avaliacao(

        nome,

        "VACINAS",

        dados

    )

# ==========================================================
# LER VACINAÇÃO
# ==========================================================

def ler_vacinacao(nome):

    """
    Lê a aba VACINAS.
    """

    mapa = {

        "antirrabica":"B2",

        "v8_v10":"B3",

        "giardia":"B4",

        "leishmaniose":"B5",

        "ultima":"B6",

        "proxima":"B7",

        "vermifugo":"B8",

        "pulgas":"B9",

        "observacoes":"B10"

    }

    return ler_avaliacao(

        nome,

        "VACINAS",

        mapa

    )

# ==========================================================
# EXISTE VACINAÇÃO
# ==========================================================

def existe_vacinacao(nome):

    """
    Verifica se existe
    a aba VACINAS.
    """

    return existe_avaliacao(

        nome,

        "VACINAS"

    )

# ==========================================================
# PRÓXIMA VACINA
# ==========================================================

def proxima_vacina(nome):

    """
    Retorna a data da próxima vacina.
    """

    dados = ler_vacinacao(

        nome

    )

    return dados.get(

        "proxima",

        ""

    )

# ==========================================================
# FIM DA PARTE 12/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 13/20
# RELATÓRIOS
# ==========================================================

# ==========================================================
# CAMINHO DO RELATÓRIO
# ==========================================================

def caminho_relatorio(nome):

    """
    Retorna o caminho do PDF
    do paciente.
    """

    nome = limpar_texto(nome)

    return PASTA_RELATORIOS / f"{nome}.pdf"

# ==========================================================
# EXISTE RELATÓRIO
# ==========================================================

def relatorio_existe(nome):

    """
    Verifica se existe PDF.
    """

    return caminho_relatorio(

        nome

    ).exists()

# ==========================================================
# LISTAR RELATÓRIOS
# ==========================================================

def listar_relatorios():

    """
    Lista todos os PDFs.
    """

    lista = []

    for arquivo in sorted(

        PASTA_RELATORIOS.glob(

            "*.pdf"

        )

    ):

        lista.append(

            arquivo.stem

        )

    return lista

# ==========================================================
# TOTAL DE RELATÓRIOS
# ==========================================================

def total_relatorios():

    """
    Quantidade de PDFs.
    """

    return len(

        listar_relatorios()

    )

# ==========================================================
# ÚLTIMO RELATÓRIO
# ==========================================================

def ultimo_relatorio():

    """
    Último PDF criado.
    """

    arquivos = sorted(

        PASTA_RELATORIOS.glob(

            "*.pdf"

        ),

        key=lambda arq: arq.stat().st_mtime,

        reverse=True

    )

    if not arquivos:

        return "-"

    return arquivos[0].stem

# ==========================================================
# EXCLUIR RELATÓRIO
# ==========================================================

def excluir_relatorio(nome):

    """
    Remove um PDF.
    """

    arquivo = caminho_relatorio(

        nome

    )

    if arquivo.exists():

        arquivo.unlink()

        return True

    return False

# ==========================================================
# RESUMO DOS RELATÓRIOS
# ==========================================================

def resumo_relatorios():

    """
    Informações rápidas.
    """

    return {

        "total": total_relatorios(),

        "ultimo": ultimo_relatorio()

    }

# ==========================================================
# FIM DA PARTE 13/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 14/20
# DASHBOARD
# ==========================================================

# ==========================================================
# TOTAL DE ESPÉCIES
# ==========================================================

def total_especies():

    """
    Retorna a quantidade de
    espécies cadastradas.
    """

    especies = set()

    for pet in listar_dados_pets():

        especie = str(

            pet.get(

                "especie",

                ""

            )

        ).strip()

        if especie:

            especies.add(

                especie

            )

    return len(

        especies

    )

# ==========================================================
# PESO MÉDIO
# ==========================================================

def peso_medio():

    """
    Calcula o peso médio
    dos pets.
    """

    pesos = []

    for pet in listar_dados_pets():

        try:

            peso = float(

                pet.get(

                    "peso",

                    0

                )

            )

            if peso > 0:

                pesos.append(

                    peso

                )

        except Exception:

            pass

    if not pesos:

        return 0

    return round(

        sum(pesos) / len(pesos),

        2

    )

# ==========================================================
# DASHBOARD
# ==========================================================

def dashboard():

    """
    Retorna os principais
    indicadores do sistema.
    """

    return {

        "total_pets": total_pets(),

        "ultimo_pet": ultimo_pet(),

        "total_relatorios": total_relatorios(),

        "ultimo_relatorio": ultimo_relatorio(),

        "especies": total_especies(),

        "peso_medio": peso_medio(),

        "por_especie": total_por_especie(),

        "por_sexo": total_por_sexo()

    }

# ==========================================================
# RESUMO
# ==========================================================

def resumo():

    """
    Resumo utilizado pela Home.
    """

    info = dashboard()

    return {

        "total": info["total_pets"],

        "ultimo": info["ultimo_pet"],

        "relatorios": info["total_relatorios"],

        "peso_medio": info["peso_medio"],

        "especies": info["especies"]

    }

# ==========================================================
# FIM DA PARTE 14/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 15/20
# BACKUP E UTILIDADES
# ==========================================================

# ==========================================================
# BACKUP DO PET
# ==========================================================

def backup_pet(nome):

    """
    Cria uma cópia de segurança
    do arquivo do pet.
    """

    origem = caminho_pet(nome)

    if not origem.exists():

        raise FileNotFoundError(

            f"Pet '{nome}' não encontrado."

        )

    data = datetime.now().strftime(

        "%Y%m%d_%H%M%S"

    )

    destino = PASTA_PETS / f"{nome}_backup_{data}.xlsx"

    copy2(

        origem,

        destino

    )

    return destino

# ==========================================================
# TAMANHO DO ARQUIVO
# ==========================================================

def tamanho_pet(nome):

    """
    Retorna o tamanho do arquivo
    em KB.
    """

    arquivo = caminho_pet(nome)

    if not arquivo.exists():

        return 0

    return round(

        arquivo.stat().st_size / 1024,

        2

    )

# ==========================================================
# DATA DE MODIFICAÇÃO
# ==========================================================

def data_modificacao(nome):

    """
    Retorna a última modificação
    do arquivo.
    """

    arquivo = caminho_pet(nome)

    if not arquivo.exists():

        return "-"

    return datetime.fromtimestamp(

        arquivo.stat().st_mtime

    ).strftime(

        "%d/%m/%Y %H:%M"

    )

# ==========================================================
# INFORMAÇÕES DO PET
# ==========================================================

def info_pet(nome):

    """
    Informações do arquivo.
    """

    return {

        "arquivo": caminho_pet(nome).name,

        "tamanho_kb": tamanho_pet(nome),

        "ultima_modificacao": data_modificacao(nome),

        "relatorio": relatorio_existe(nome)

    }

# ==========================================================
# VERIFICAÇÃO GERAL
# ==========================================================

def sistema_ok():

    """
    Verifica se o sistema
    está pronto para uso.
    """

    return {

        "modelo": MODELO_PET.exists(),

        "pets": PASTA_PETS.exists(),

        "relatorios": PASTA_RELATORIOS.exists()

    }

# ==========================================================
# FIM DA PARTE 15/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 16/20
# ESTATÍSTICAS
# ==========================================================

# ==========================================================
# ESPÉCIES CADASTRADAS
# ==========================================================

def especies_cadastradas():

    """
    Lista todas as espécies cadastradas.
    """

    especies = set()

    for pet in listar_dados_pets():

        especie = str(

            pet.get(

                "especie",

                ""

            )

        ).strip()

        if especie:

            especies.add(

                especie

            )

    return sorted(

        especies

    )

# ==========================================================
# TUTORES CADASTRADOS
# ==========================================================

def tutores_cadastrados():

    """
    Lista todos os tutores.
    """

    tutores = set()

    for pet in listar_dados_pets():

        tutor = str(

            pet.get(

                "tutor",

                ""

            )

        ).strip()

        if tutor:

            tutores.add(

                tutor

            )

    return sorted(

        tutores

    )

# ==========================================================
# TOTAL DE TUTORES
# ==========================================================

def total_tutores():

    """
    Quantidade de tutores.
    """

    return len(

        tutores_cadastrados()

    )

# ==========================================================
# PESO TOTAL
# ==========================================================

def peso_total():

    """
    Soma o peso de todos os pets.
    """

    total = 0

    for pet in listar_dados_pets():

        try:

            total += float(

                pet.get(

                    "peso",

                    0

                )

            )

        except Exception:

            pass

    return round(

        total,

        2

    )

# ==========================================================
# RESUMO GERAL
# ==========================================================

def resumo_geral():

    """
    Resumo utilizado pelo sistema.
    """

    return {

        "pets": total_pets(),

        "tutores": total_tutores(),

        "especies": total_especies(),

        "peso_total": peso_total(),

        "peso_medio": peso_medio(),

        "relatorios": total_relatorios()

    }

# ==========================================================
# FIM DA PARTE 16/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 17/20
# PESQUISA E FILTROS
# ==========================================================

# ==========================================================
# PESQUISAR POR NOME
# ==========================================================

def pesquisar_por_nome(texto):

    """
    Pesquisa pets pelo nome.
    """

    texto = limpar_texto(texto).lower()

    return [

        nome

        for nome in listar_pets()

        if texto in nome.lower()

    ]

# ==========================================================
# FILTRAR POR ESPÉCIE
# ==========================================================

def filtrar_por_especie(especie):

    """
    Retorna todos os pets
    da espécie informada.
    """

    resultado = []

    especie = limpar_texto(especie).lower()

    for pet in listar_dados_pets():

        if str(

            pet.get(

                "especie",

                ""

            )

        ).lower() == especie:

            resultado.append(

                pet

            )

    return resultado

# ==========================================================
# FILTRAR POR TUTOR
# ==========================================================

def filtrar_por_tutor(tutor):

    """
    Retorna todos os pets
    do tutor informado.
    """

    resultado = []

    tutor = limpar_texto(tutor).lower()

    for pet in listar_dados_pets():

        if tutor in str(

            pet.get(

                "tutor",

                ""

            )

        ).lower():

            resultado.append(

                pet

            )

    return resultado

# ==========================================================
# FILTRAR POR SEXO
# ==========================================================

def filtrar_por_sexo(sexo):

    """
    Filtra Macho ou Fêmea.
    """

    resultado = []

    sexo = limpar_texto(sexo).lower()

    for pet in listar_dados_pets():

        if str(

            pet.get(

                "sexo",

                ""

            )

        ).lower() == sexo:

            resultado.append(

                pet

            )

    return resultado

# ==========================================================
# PESQUISA GERAL
# ==========================================================

def pesquisar(texto):

    """
    Pesquisa por nome,
    tutor ou espécie.
    """

    texto = limpar_texto(texto).lower()

    resultado = []

    for pet in listar_dados_pets():

        if (

            texto in str(

                pet.get(

                    "nome",

                    ""

                )

            ).lower()

            or

            texto in str(

                pet.get(

                    "tutor",

                    ""

                )

            ).lower()

            or

            texto in str(

                pet.get(

                    "especie",

                    ""

                )

            ).lower()

        ):

            resultado.append(

                pet

            )

    return resultado

# ==========================================================
# FIM DA PARTE 17/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 18/20
# MANUTENÇÃO E EXPORTAÇÃO
# ==========================================================

# ==========================================================
# LISTA DE ARQUIVOS
# ==========================================================

def listar_arquivos():

    """
    Retorna todos os arquivos Excel
    da pasta pets.
    """

    return sorted(

        PASTA_PETS.glob(

            "*.xlsx"

        )

    )

# ==========================================================
# TAMANHO TOTAL
# ==========================================================

def tamanho_total():

    """
    Tamanho total ocupado
    pelos cadastros.
    """

    total = 0

    for arquivo in listar_arquivos():

        total += arquivo.stat().st_size

    return round(

        total / 1024,

        2

    )

# ==========================================================
# EXPORTAR LISTA
# ==========================================================

def exportar_lista():

    """
    Exporta somente os nomes
    dos pacientes.
    """

    return listar_pets()

# ==========================================================
# PACIENTES SEM RELATÓRIO
# ==========================================================

def pets_sem_relatorio():

    """
    Retorna os pets que ainda
    não possuem PDF.
    """

    lista = []

    for nome in listar_pets():

        if not relatorio_existe(nome):

            lista.append(

                nome

            )

    return lista

# ==========================================================
# PACIENTES COM RELATÓRIO
# ==========================================================

def pets_com_relatorio():

    """
    Retorna os pets que possuem PDF.
    """

    lista = []

    for nome in listar_pets():

        if relatorio_existe(nome):

            lista.append(

                nome

            )

    return lista

# ==========================================================
# LIMPEZA DOS BACKUPS
# ==========================================================

def limpar_backups():

    """
    Remove arquivos de backup.
    """

    quantidade = 0

    for arquivo in PASTA_PETS.glob(

        "*_backup_*.xlsx"

    ):

        arquivo.unlink()

        quantidade += 1

    return quantidade

# ==========================================================
# ESPAÇO LIVRE
# ==========================================================

def espaco_utilizado():

    """
    Retorna espaço ocupado
    pela pasta pets.
    """

    return {

        "arquivos": total_pets(),

        "tamanho_kb": tamanho_total()

    }

# ==========================================================
# FIM DA PARTE 18/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 19/20
# VERIFICAÇÃO DO SISTEMA
# ==========================================================

# ==========================================================
# VERIFICAR ESTRUTURA
# ==========================================================

def verificar_estrutura():

    """
    Verifica se a estrutura básica
    do sistema está correta.
    """

    return {

        "modelo": MODELO_PET.exists(),

        "pasta_pets": PASTA_PETS.exists(),

        "pasta_relatorios": PASTA_RELATORIOS.exists()

    }

# ==========================================================
# VERIFICAR PET
# ==========================================================

def verificar_pet(nome):

    """
    Verifica se o arquivo do pet
    pode ser aberto.
    """

    try:

        wb = abrir_planilha(nome)

        wb.close()

        return True

    except Exception:

        return False

# ==========================================================
# VERIFICAR TODOS OS PETS
# ==========================================================

def verificar_pets():

    """
    Verifica todos os arquivos
    da pasta pets.
    """

    resultado = {}

    for nome in listar_pets():

        resultado[nome] = verificar_pet(nome)

    return resultado

# ==========================================================
# PETS INVÁLIDOS
# ==========================================================

def pets_invalidos():

    """
    Lista arquivos com problemas.
    """

    lista = []

    for nome, ok in verificar_pets().items():

        if not ok:

            lista.append(nome)

    return lista

# ==========================================================
# SAÚDE DO SISTEMA
# ==========================================================

def sistema_saudavel():

    """
    Retorna True quando o sistema
    estiver totalmente funcional.
    """

    estrutura = verificar_estrutura()

    if not all(estrutura.values()):

        return False

    if pets_invalidos():

        return False

    return True

# ==========================================================
# DIAGNÓSTICO
# ==========================================================

def diagnostico():

    """
    Retorna um resumo do sistema.
    """

    return {

        "sistema": sistema_saudavel(),

        "estrutura": verificar_estrutura(),

        "pets": total_pets(),

        "relatorios": total_relatorios(),

        "pets_invalidos": pets_invalidos()

    }

# ==========================================================
# FIM DA PARTE 19/20
# ==========================================================
# ==========================================================
# DINOPETANIMAL PRO 5.0
# excel.py
#
# PARTE 20/20
# FINALIZAÇÃO
# ==========================================================

# ==========================================================
# INFORMAÇÕES DO SISTEMA
# ==========================================================

VERSAO = "5.0"

SISTEMA = "DinoPetAnimal PRO"

AUTOR = "Professor Dinaldo Guedes"

# ==========================================================
# SOBRE O SISTEMA
# ==========================================================

def sobre():

    """
    Informações do sistema.
    """

    return {

        "sistema": SISTEMA,

        "versao": VERSAO,

        "autor": AUTOR,

        "data": agora_texto()

    }

# ==========================================================
# RESUMO FINAL
# ==========================================================

def resumo_sistema():

    """
    Resumo geral utilizado
    pela Home e Dashboard.
    """

    return {

        "sistema": SISTEMA,

        "versao": VERSAO,

        "pets": total_pets(),

        "ultimo_pet": ultimo_pet(),

        "especies": total_especies(),

        "tutores": total_tutores(),

        "relatorios": total_relatorios(),

        "peso_medio": peso_medio(),

        "sistema_ok": sistema_saudavel()

    }

# ==========================================================
# TESTE DO SISTEMA
# ==========================================================

def testar():

    """
    Executa uma verificação rápida.
    """

    try:

        resumo_sistema()

        return True

    except Exception:

        return False

# ==========================================================
# EXPORTAÇÃO
# ==========================================================

__all__ = [

    "criar_pet",

    "abrir_pet",

    "ler_cadastro",

    "salvar_cadastro",

    "atualizar_pet",

    "renomear_pet",

    "excluir_pet",

    "listar_pets",

    "pesquisar",

    "dashboard",

    "resumo",

    "resumo_geral",

    "resumo_sistema",

    "diagnostico",

    "sobre"

]

# ==========================================================
# FIM
# ==========================================================