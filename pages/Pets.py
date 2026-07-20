from pathlib import Path
from datetime import date, datetime
import re

import streamlit as st
from openpyxl import load_workbook

# ==========================================================
# DINOPETANIMAL PRO
# PÁGINA: PETS
# ==========================================================


# ==========================================================
# CAMINHOS
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_PETS = BASE_DIR / "pets"
ARQUIVO_CSS = BASE_DIR / "styles.css"

PASTA_PETS.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CSS
# ==========================================================

if ARQUIVO_CSS.exists() and ARQUIVO_CSS.stat().st_size > 0:
    st.markdown(
        f"<style>{ARQUIVO_CSS.read_text(encoding='utf-8')}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    .pets-topo {
        background: linear-gradient(135deg, #fff7ed, #ffedd5);
        border: 1px solid #fdba74;
        border-radius: 18px;
        padding: 22px 26px;
        margin-bottom: 22px;
    }

    .pets-topo h1 {
        margin: 0;
        color: #1f2937;
        font-size: 2.2rem;
    }

    .pets-topo p {
        margin: 8px 0 0 0;
        color: #475569;
        font-size: 1.05rem;
    }

    .pet-card {
        background: #ffffff;
        border: 1px solid #fed7aa;
        border-radius: 16px;
        padding: 18px 20px;
        margin-bottom: 14px;
        box-shadow: 0 2px 10px rgba(0, 0, 0, .04);
    }

    .pet-card h3 {
        margin: 0 0 8px 0;
        color: #7c2d12;
    }

    .pet-card p {
        margin: 4px 0;
        color: #475569;
    }

    div[data-testid="stForm"] {
        border: 1px solid #fed7aa;
        border-radius: 18px;
        padding: 22px;
        background: #ffffff;
    }

    .stButton > button,
    .stDownloadButton > button,
    .stFormSubmitButton > button {
        border-radius: 10px;
        min-height: 44px;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# FUNÇÕES
# ==========================================================

def texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def listar_arquivos_pets():
    return sorted(
        [
            arquivo
            for arquivo in PASTA_PETS.glob("*.xlsx")
            if not arquivo.name.startswith("~$")
        ],
        key=lambda caminho: caminho.stem.casefold(),
    )


def localizar_linha(ws, rotulo: str):
    procurado = texto_limpo(rotulo).casefold()

    for linha in range(1, ws.max_row + 1):
        atual = texto_limpo(ws.cell(linha, 1).value).casefold()
        if atual == procurado:
            return linha

    return None


def ler_campo(ws, rotulo: str, padrao=""):
    linha = localizar_linha(ws, rotulo)
    if linha is None:
        return padrao

    valor = ws.cell(linha, 2).value
    return padrao if valor is None else valor


def escrever_campo(ws, rotulo: str, valor):
    linha = localizar_linha(ws, rotulo)

    if linha is None:
        raise KeyError(f"Campo não encontrado no modelo: {rotulo}")

    ws.cell(linha, 2).value = valor


def carregar_pet(arquivo: Path) -> dict:
    wb = load_workbook(arquivo, data_only=False)

    if "Cadastro" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba 'Cadastro' não foi encontrada.")

    ws = wb["Cadastro"]

    dados = {
        "arquivo": arquivo,
        "nome": texto_limpo(ler_campo(ws, "Nome do Pet", arquivo.stem)),
        "microchip": texto_limpo(ler_campo(ws, "Microchip")),
        "especie": texto_limpo(ler_campo(ws, "Espécie")),
        "raca": texto_limpo(ler_campo(ws, "Raça")),
        "sexo": texto_limpo(ler_campo(ws, "Sexo")),
        "castrado": texto_limpo(ler_campo(ws, "Castrado")),
        "nascimento": ler_campo(ws, "Data de Nascimento"),
        "idade": texto_limpo(ler_campo(ws, "Idade")),
        "cor": texto_limpo(ler_campo(ws, "Cor")),
        "peso": ler_campo(ws, "Peso (kg)", 0),
        "tutor": texto_limpo(ler_campo(ws, "Tutor")),
        "cpf": texto_limpo(ler_campo(ws, "CPF")),
        "telefone": texto_limpo(ler_campo(ws, "Telefone")),
        "email": texto_limpo(ler_campo(ws, "E-mail")),
        "endereco": texto_limpo(ler_campo(ws, "Endereço")),
        "veterinario": texto_limpo(ler_campo(ws, "Veterinário")),
        "crmv": texto_limpo(ler_campo(ws, "CRMV")),
        "observacoes": texto_limpo(ler_campo(ws, "Observações")),
    }

    wb.close()
    return dados


def converter_data(valor):
    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = texto_limpo(valor)

    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            pass

    return date.today()


def calcular_idade(data_nascimento: date) -> str:
    hoje = date.today()

    if data_nascimento > hoje:
        return ""

    anos = hoje.year - data_nascimento.year
    meses = hoje.month - data_nascimento.month

    if hoje.day < data_nascimento.day:
        meses -= 1

    if meses < 0:
        anos -= 1
        meses += 12

    if anos <= 0:
        return f"{meses} mês" if meses == 1 else f"{meses} meses"

    if meses == 0:
        return f"{anos} ano" if anos == 1 else f"{anos} anos"

    texto_anos = f"{anos} ano" if anos == 1 else f"{anos} anos"
    texto_meses = f"{meses} mês" if meses == 1 else f"{meses} meses"
    return f"{texto_anos} e {texto_meses}"


def nome_seguro_arquivo(nome: str) -> str:
    nome = texto_limpo(nome)
    nome = re.sub(r'[\\/:*?"<>|]', "", nome)
    return nome.rstrip(". ")


def salvar_alteracoes(arquivo_atual: Path, dados: dict) -> Path:
    nome_novo = nome_seguro_arquivo(dados["nome"])

    if len(nome_novo) < 2:
        raise ValueError("Informe um nome válido para o pet.")

    novo_caminho = PASTA_PETS / f"{nome_novo}.xlsx"

    if novo_caminho != arquivo_atual and novo_caminho.exists():
        raise FileExistsError(
            f"Já existe um pet chamado '{nome_novo}'."
        )

    wb = load_workbook(arquivo_atual)
    ws = wb["Cadastro"]

    campos = {
        "Nome do Pet": nome_novo,
        "Microchip": dados["microchip"],
        "Espécie": dados["especie"],
        "Raça": dados["raca"],
        "Sexo": dados["sexo"],
        "Castrado": dados["castrado"],
        "Data de Nascimento": dados["nascimento"].strftime("%d/%m/%Y"),
        "Idade": calcular_idade(dados["nascimento"]),
        "Cor": dados["cor"],
        "Peso (kg)": float(dados["peso"]),
        "Tutor": dados["tutor"],
        "CPF": dados["cpf"],
        "Telefone": dados["telefone"],
        "E-mail": dados["email"],
        "Endereço": dados["endereco"],
        "Veterinário": dados["veterinario"],
        "CRMV": dados["crmv"],
        "Observações": dados["observacoes"],
    }

    for rotulo, valor in campos.items():
        escrever_campo(ws, rotulo, valor)

    if "Oculto" in wb.sheetnames:
        ws_oculto = wb["Oculto"]
        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        if ws_oculto.max_row >= 4:
            ws_oculto["A4"] = agora

    wb.save(arquivo_atual)
    wb.close()

    if novo_caminho != arquivo_atual:
        arquivo_atual.rename(novo_caminho)

    return novo_caminho


def excluir_pet(arquivo: Path):
    if not arquivo.exists():
        raise FileNotFoundError("O arquivo do pet não foi encontrado.")

    arquivo.unlink()


def valor_opcao(valor_atual: str, opcoes: list[str]) -> int:
    try:
        return opcoes.index(valor_atual)
    except ValueError:
        return 0


# ==========================================================
# CABEÇALHO
# ==========================================================

st.markdown(
    """
    <div class="pets-topo">
        <h1>🐾 Pets cadastrados</h1>
        <p>Consulte, edite, baixe ou exclua os cadastros dos animais.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

arquivos = listar_arquivos_pets()

col_metrica1, col_metrica2, col_metrica3 = st.columns(3)

with col_metrica1:
    st.metric("🐕 Total de pets", len(arquivos))

with col_metrica2:
    especies = set()
    for arquivo in arquivos:
        try:
            especie = carregar_pet(arquivo).get("especie")
            if especie:
                especies.add(especie)
        except Exception:
            pass
    st.metric("🦴 Espécies", len(especies))

with col_metrica3:
    st.metric(
        "📄 Arquivos Excel",
        len(arquivos),
    )

st.divider()

# ==========================================================
# SEM CADASTROS
# ==========================================================

if not arquivos:
    st.info(
        "Nenhum pet foi cadastrado até o momento. "
        "Use a página Cadastro para adicionar o primeiro animal."
    )

    if st.button(
        "➕ Ir para Cadastro",
        type="primary",
        use_container_width=True,
    ):
        st.switch_page("pages/Cadastro.py")

    st.stop()

# ==========================================================
# PESQUISA E SELEÇÃO
# ==========================================================

col_busca, col_ordenacao = st.columns([3, 1])

with col_busca:
    pesquisa = st.text_input(
        "🔎 Pesquisar pet",
        placeholder="Digite nome, tutor, espécie ou raça",
    )

with col_ordenacao:
    ordem = st.selectbox(
        "Ordenar",
        ["Nome A-Z", "Nome Z-A", "Mais recente"],
    )

dados_pets = []

for arquivo in arquivos:
    try:
        dados_pets.append(carregar_pet(arquivo))
    except Exception as erro:
        st.warning(
            f"Não foi possível ler o arquivo {arquivo.name}: {erro}"
        )

texto_pesquisa = texto_limpo(pesquisa).casefold()

if texto_pesquisa:
    dados_pets = [
        pet
        for pet in dados_pets
        if texto_pesquisa
        in " ".join(
            [
                pet.get("nome", ""),
                pet.get("tutor", ""),
                pet.get("especie", ""),
                pet.get("raca", ""),
                pet.get("telefone", ""),
            ]
        ).casefold()
    ]

if ordem == "Nome A-Z":
    dados_pets.sort(key=lambda item: item["nome"].casefold())
elif ordem == "Nome Z-A":
    dados_pets.sort(
        key=lambda item: item["nome"].casefold(),
        reverse=True,
    )
else:
    dados_pets.sort(
        key=lambda item: item["arquivo"].stat().st_mtime,
        reverse=True,
    )

if not dados_pets:
    st.warning("Nenhum pet corresponde à pesquisa informada.")
    st.stop()

nomes_exibicao = [
    f"{pet['nome']} — {pet['especie'] or 'Espécie não informada'}"
    for pet in dados_pets
]

indice_padrao = 0
ultimo_pet = st.session_state.get("ultimo_pet_cadastrado")

if ultimo_pet:
    for indice, pet in enumerate(dados_pets):
        if pet["nome"].casefold() == str(ultimo_pet).casefold():
            indice_padrao = indice
            break

selecionado_exibicao = st.selectbox(
    "Selecione um pet para abrir",
    nomes_exibicao,
    index=indice_padrao,
)

indice_selecionado = nomes_exibicao.index(selecionado_exibicao)
pet = dados_pets[indice_selecionado]
arquivo_pet = pet["arquivo"]

st.session_state["pet_selecionado"] = pet["nome"]

# ==========================================================
# RESUMO
# ==========================================================

st.markdown(
    f"""
    <div class="pet-card">
        <h3>🐾 {pet["nome"]}</h3>
        <p><b>Espécie:</b> {pet["especie"] or "-"}</p>
        <p><b>Raça:</b> {pet["raca"] or "-"}</p>
        <p><b>Tutor:</b> {pet["tutor"] or "-"}</p>
        <p><b>Telefone:</b> {pet["telefone"] or "-"}</p>
    </div>
    """,
    unsafe_allow_html=True,
)

col_acao1, col_acao2, col_acao3 = st.columns(3)

with col_acao1:
    with open(arquivo_pet, "rb") as arquivo_download:
        st.download_button(
            "⬇️ Baixar planilha",
            data=arquivo_download.read(),
            file_name=arquivo_pet.name,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
        )

with col_acao2:
    if st.button(
        "🩺 Abrir avaliação",
        use_container_width=True,
    ):
        st.session_state["pet_selecionado"] = pet["nome"]
        st.switch_page("pages/Avaliacao.py")

with col_acao3:
    if st.button(
        "🔄 Atualizar lista",
        use_container_width=True,
    ):
        st.rerun()

st.divider()

# ==========================================================
# EDIÇÃO
# ==========================================================

with st.expander(
    f"✏️ Editar cadastro de {pet['nome']}",
    expanded=True,
):
    nascimento_atual = converter_data(pet["nascimento"])

    especies_opcoes = [
        "Cão",
        "Gato",
        "Ave",
        "Coelho",
        "Roedor",
        "Réptil",
        "Outro",
    ]
    if pet["especie"] and pet["especie"] not in especies_opcoes:
        especies_opcoes.append(pet["especie"])

    sexos_opcoes = ["Macho", "Fêmea", "Não informado"]
    if pet["sexo"] and pet["sexo"] not in sexos_opcoes:
        sexos_opcoes.append(pet["sexo"])

    castrado_opcoes = ["Não", "Sim", "Não informado"]
    if pet["castrado"] and pet["castrado"] not in castrado_opcoes:
        castrado_opcoes.append(pet["castrado"])

    with st.form(
        f"form_editar_{arquivo_pet.stem}",
        clear_on_submit=False,
    ):
        st.subheader("🐾 Dados do animal")

        col1, col2, col3 = st.columns(3)

        with col1:
            nome = st.text_input(
                "Nome do pet *",
                value=pet["nome"],
            )

            especie = st.selectbox(
                "Espécie",
                especies_opcoes,
                index=valor_opcao(
                    pet["especie"],
                    especies_opcoes,
                ),
            )

            sexo = st.selectbox(
                "Sexo",
                sexos_opcoes,
                index=valor_opcao(
                    pet["sexo"],
                    sexos_opcoes,
                ),
            )

        with col2:
            raca = st.text_input(
                "Raça",
                value=pet["raca"],
            )

            cor = st.text_input(
                "Cor",
                value=pet["cor"],
            )

            castrado = st.selectbox(
                "Castrado",
                castrado_opcoes,
                index=valor_opcao(
                    pet["castrado"],
                    castrado_opcoes,
                ),
            )

        with col3:
            nascimento = st.date_input(
                "Data de nascimento",
                value=nascimento_atual,
                max_value=date.today(),
                format="DD/MM/YYYY",
            )

            st.text_input(
                "Idade calculada",
                value=calcular_idade(nascimento),
                disabled=True,
            )

            try:
                peso_atual = float(pet["peso"] or 0)
            except (TypeError, ValueError):
                peso_atual = 0.0

            peso = st.number_input(
                "Peso (kg)",
                min_value=0.0,
                max_value=500.0,
                value=peso_atual,
                step=0.1,
                format="%.2f",
            )

        microchip = st.text_input(
            "Microchip",
            value=pet["microchip"],
        )

        st.divider()
        st.subheader("👤 Dados do tutor")

        col4, col5 = st.columns(2)

        with col4:
            tutor = st.text_input(
                "Tutor *",
                value=pet["tutor"],
            )

            cpf = st.text_input(
                "CPF",
                value=pet["cpf"],
            )

            telefone = st.text_input(
                "Telefone/WhatsApp",
                value=pet["telefone"],
            )

        with col5:
            email = st.text_input(
                "E-mail",
                value=pet["email"],
            )

            endereco = st.text_area(
                "Endereço",
                value=pet["endereco"],
                height=118,
            )

        st.divider()
        st.subheader("🩺 Atendimento veterinário")

        col6, col7 = st.columns(2)

        with col6:
            veterinario = st.text_input(
                "Veterinário",
                value=pet["veterinario"],
            )

        with col7:
            crmv = st.text_input(
                "CRMV",
                value=pet["crmv"],
            )

        observacoes = st.text_area(
            "Observações",
            value=pet["observacoes"],
            height=130,
        )

        salvar = st.form_submit_button(
            "💾 Salvar alterações",
            type="primary",
            use_container_width=True,
        )

    if salvar:
        erros = []

        if len(texto_limpo(nome)) < 2:
            erros.append("Informe o nome do pet.")

        if len(texto_limpo(tutor)) < 2:
            erros.append("Informe o nome do tutor.")

        if peso <= 0:
            erros.append("Informe um peso maior que zero.")

        if email and (
            "@" not in email or "." not in email.split("@")[-1]
        ):
            erros.append("O e-mail informado parece inválido.")

        if erros:
            for erro in erros:
                st.error(f"❌ {erro}")
        else:
            dados_atualizados = {
                "nome": texto_limpo(nome),
                "microchip": texto_limpo(microchip),
                "especie": texto_limpo(especie),
                "raca": texto_limpo(raca),
                "sexo": texto_limpo(sexo),
                "castrado": texto_limpo(castrado),
                "nascimento": nascimento,
                "cor": texto_limpo(cor),
                "peso": float(peso),
                "tutor": texto_limpo(tutor),
                "cpf": texto_limpo(cpf),
                "telefone": texto_limpo(telefone),
                "email": texto_limpo(email),
                "endereco": texto_limpo(endereco),
                "veterinario": texto_limpo(veterinario),
                "crmv": texto_limpo(crmv),
                "observacoes": texto_limpo(observacoes),
            }

            try:
                novo_arquivo = salvar_alteracoes(
                    arquivo_pet,
                    dados_atualizados,
                )

                st.session_state["pet_selecionado"] = (
                    dados_atualizados["nome"]
                )
                st.session_state["ultimo_pet_cadastrado"] = (
                    dados_atualizados["nome"]
                )

                st.success(
                    f"✅ Cadastro de **{dados_atualizados['nome']}** "
                    "atualizado com sucesso."
                )
                st.info(f"📁 Arquivo: pets/{novo_arquivo.name}")

                st.rerun()

            except Exception as erro:
                st.error("❌ Não foi possível salvar as alterações.")
                st.exception(erro)

# ==========================================================
# EXCLUSÃO
# ==========================================================

st.divider()

with st.expander("🗑️ Excluir pet", expanded=False):
    st.warning(
        "Esta ação excluirá definitivamente a planilha individual "
        "do pet. Não será possível desfazer."
    )

    confirmar = st.checkbox(
        f"Confirmo que desejo excluir {pet['nome']}.",
        key=f"confirmar_exclusao_{arquivo_pet.stem}",
    )

    texto_confirmacao = st.text_input(
        "Digite EXCLUIR para confirmar",
        key=f"texto_exclusao_{arquivo_pet.stem}",
    )

    if st.button(
        "🗑️ Excluir definitivamente",
        type="primary",
        use_container_width=True,
        disabled=not (
            confirmar
            and texto_confirmacao.strip().upper() == "EXCLUIR"
        ),
    ):
        try:
            nome_excluido = pet["nome"]
            excluir_pet(arquivo_pet)

            st.session_state.pop("pet_selecionado", None)
            st.session_state.pop("ultimo_pet_cadastrado", None)

            st.success(
                f"✅ O cadastro de **{nome_excluido}** foi excluído."
            )
            st.rerun()

        except Exception as erro:
            st.error("❌ Não foi possível excluir o pet.")
            st.exception(erro)

# ==========================================================
# RODAPÉ
# ==========================================================

st.divider()
st.caption(
    "DinoPetAnimal PRO • Gestão individual de pets em Excel"
)
