from pathlib import Path
from collections import Counter

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# ==========================================================
# DINOPETANIMAL PRO
# PÁGINA: DASHBOARD
# ==========================================================


# ==========================================================
# CAMINHOS
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_PETS = BASE_DIR / "pets"
ARQUIVO_CSS = BASE_DIR / "styles.css"

PASTA_PETS.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CSS
# ==========================================================

if ARQUIVO_CSS.exists() and ARQUIVO_CSS.stat().st_size > 0:
    st.markdown(
        f"<style>{ARQUIVO_CSS.read_text(encoding='utf-8')}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    .dashboard-topo {
        background: linear-gradient(135deg, #fff7ed, #ffedd5);
        border: 1px solid #fdba74;
        border-radius: 18px;
        padding: 22px 26px;
        margin-bottom: 22px;
    }

    .dashboard-topo h1 {
        margin: 0;
        color: #1f2937;
        font-size: 2.2rem;
    }

    .dashboard-topo p {
        margin: 8px 0 0 0;
        color: #475569;
        font-size: 1.05rem;
    }

    .painel-card {
        background: #ffffff;
        border: 1px solid #fed7aa;
        border-radius: 16px;
        padding: 18px 20px;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(0, 0, 0, .04);
    }

    .painel-card h3 {
        margin: 0 0 8px 0;
        color: #7c2d12;
    }

    .stButton > button {
        border-radius: 10px;
        min-height: 44px;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# FUNÇÕES
# ==========================================================

def texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def listar_arquivos_pets():
    return sorted(
        [
            arquivo
            for arquivo in PASTA_PETS.glob("*.xlsx")
            if not arquivo.name.startswith("~$")
        ],
        key=lambda arquivo: arquivo.stem.casefold(),
    )


def localizar_linha(ws, rotulo: str):
    procurado = texto_limpo(rotulo).casefold()

    for linha in range(1, ws.max_row + 1):
        atual = texto_limpo(ws.cell(linha, 1).value).casefold()
        if atual == procurado:
            return linha

    return None


def ler_campo(ws, rotulo: str, coluna=2, padrao=""):
    linha = localizar_linha(ws, rotulo)

    if linha is None:
        return padrao

    valor = ws.cell(linha, coluna).value
    return padrao if valor is None else valor


def converter_numero(valor, padrao=0.0):
    try:
        return float(valor)
    except (TypeError, ValueError):
        return padrao


def carregar_pet(arquivo: Path) -> dict:
    wb = load_workbook(arquivo, data_only=True)

    if "Cadastro" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba Cadastro não foi encontrada.")

    ws_cadastro = wb["Cadastro"]

    dados = {
        "arquivo": arquivo,
        "nome": texto_limpo(
            ler_campo(ws_cadastro, "Nome do Pet", 2, arquivo.stem)
        ),
        "especie": texto_limpo(
            ler_campo(ws_cadastro, "Espécie")
        ),
        "raca": texto_limpo(
            ler_campo(ws_cadastro, "Raça")
        ),
        "sexo": texto_limpo(
            ler_campo(ws_cadastro, "Sexo")
        ),
        "castrado": texto_limpo(
            ler_campo(ws_cadastro, "Castrado")
        ),
        "peso": converter_numero(
            ler_campo(ws_cadastro, "Peso (kg)", 2, 0)
        ),
        "tutor": texto_limpo(
            ler_campo(ws_cadastro, "Tutor")
        ),
        "veterinario": texto_limpo(
            ler_campo(ws_cadastro, "Veterinário")
        ),
        "score": 0.0,
        "media": 0.0,
        "classificacao": "Sem avaliação",
        "data_avaliacao": "",
    }

    if "Resultados" in wb.sheetnames:
        ws_resultados = wb["Resultados"]

        dados["score"] = converter_numero(
            ler_campo(ws_resultados, "Score Geral", 2, 0)
        )
        dados["media"] = converter_numero(
            ler_campo(ws_resultados, "Média", 2, 0)
        )
        dados["classificacao"] = texto_limpo(
            ler_campo(
                ws_resultados,
                "Classificação",
                2,
                "Sem avaliação",
            )
        ) or "Sem avaliação"
        dados["data_avaliacao"] = texto_limpo(
            ler_campo(ws_resultados, "Data da Avaliação")
        )

    wb.close()
    return dados


def resumo_geral(pets: list[dict]) -> dict:
    total = len(pets)
    avaliados = sum(
        1
        for pet in pets
        if pet["classificacao"] != "Sem avaliação"
    )

    medias_validas = [
        pet["media"]
        for pet in pets
        if pet["classificacao"] != "Sem avaliação"
    ]

    media_geral = (
        sum(medias_validas) / len(medias_validas)
        if medias_validas
        else 0.0
    )

    especies = {
        pet["especie"]
        for pet in pets
        if pet["especie"]
    }

    return {
        "total": total,
        "avaliados": avaliados,
        "pendentes": total - avaliados,
        "media_geral": round(media_geral, 2),
        "especies": len(especies),
    }


def grafico_barras(contagem: dict, titulo: str, eixo_y: str):
    if not contagem:
        st.info("Ainda não existem dados suficientes para este gráfico.")
        return

    nomes = list(contagem.keys())
    valores = list(contagem.values())

    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar(nomes, valores)
    ax.set_title(titulo)
    ax.set_ylabel(eixo_y)
    ax.set_xlabel("")
    ax.tick_params(axis="x", rotation=30)
    ax.grid(axis="y", alpha=0.25)

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


def grafico_medias(pets_avaliados: list[dict]):
    if not pets_avaliados:
        st.info("Nenhum pet possui avaliação salva.")
        return

    ordenados = sorted(
        pets_avaliados,
        key=lambda item: item["media"],
        reverse=True,
    )

    nomes = [item["nome"] for item in ordenados]
    medias = [item["media"] for item in ordenados]

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(nomes, medias)
    ax.set_title("Média da avaliação por pet")
    ax.set_ylabel("Média")
    ax.set_ylim(0, 10)
    ax.tick_params(axis="x", rotation=35)
    ax.grid(axis="y", alpha=0.25)

    for indice, valor in enumerate(medias):
        ax.text(
            indice,
            valor + 0.12,
            f"{valor:.1f}",
            ha="center",
            fontsize=9,
        )

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


# ==========================================================
# CABEÇALHO
# ==========================================================

st.markdown(
    """
    <div class="dashboard-topo">
        <h1>📊 Dashboard</h1>
        <p>Visão geral dos pets, avaliações e indicadores do sistema.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

arquivos = listar_arquivos_pets()
pets = []

for arquivo in arquivos:
    try:
        pets.append(carregar_pet(arquivo))
    except Exception as erro:
        st.warning(
            f"Não foi possível carregar o arquivo {arquivo.name}: {erro}"
        )

resumo = resumo_geral(pets)

# ==========================================================
# MÉTRICAS
# ==========================================================

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.metric("🐾 Total de pets", resumo["total"])

with col2:
    st.metric("🩺 Avaliados", resumo["avaliados"])

with col3:
    st.metric("⏳ Pendentes", resumo["pendentes"])

with col4:
    st.metric("📈 Média geral", resumo["media_geral"])

with col5:
    st.metric("🦴 Espécies", resumo["especies"])

st.divider()

# ==========================================================
# SEM DADOS
# ==========================================================

if not pets:
    st.info(
        "O Dashboard ainda não possui dados. Cadastre o primeiro pet "
        "para começar a acompanhar os indicadores."
    )

    if st.button(
        "➕ Cadastrar primeiro pet",
        type="primary",
        use_container_width=True,
    ):
        st.switch_page("pages/Cadastro.py")

    st.stop()

# ==========================================================
# FILTROS
# ==========================================================

st.subheader("🔎 Filtros")

col_f1, col_f2, col_f3 = st.columns(3)

especies_disponiveis = sorted(
    {pet["especie"] for pet in pets if pet["especie"]}
)
classificacoes_disponiveis = sorted(
    {pet["classificacao"] for pet in pets if pet["classificacao"]}
)
sexos_disponiveis = sorted(
    {pet["sexo"] for pet in pets if pet["sexo"]}
)

with col_f1:
    filtro_especie = st.selectbox(
        "Espécie",
        ["Todas"] + especies_disponiveis,
    )

with col_f2:
    filtro_classificacao = st.selectbox(
        "Classificação",
        ["Todas"] + classificacoes_disponiveis,
    )

with col_f3:
    filtro_sexo = st.selectbox(
        "Sexo",
        ["Todos"] + sexos_disponiveis,
    )

pets_filtrados = pets

if filtro_especie != "Todas":
    pets_filtrados = [
        pet
        for pet in pets_filtrados
        if pet["especie"] == filtro_especie
    ]

if filtro_classificacao != "Todas":
    pets_filtrados = [
        pet
        for pet in pets_filtrados
        if pet["classificacao"] == filtro_classificacao
    ]

if filtro_sexo != "Todos":
    pets_filtrados = [
        pet
        for pet in pets_filtrados
        if pet["sexo"] == filtro_sexo
    ]

st.caption(
    f"Mostrando {len(pets_filtrados)} de {len(pets)} pets cadastrados."
)

st.divider()

# ==========================================================
# GRÁFICOS
# ==========================================================

col_g1, col_g2 = st.columns(2)

with col_g1:
    st.subheader("🐕 Pets por espécie")

    contagem_especies = Counter(
        pet["especie"] or "Não informado"
        for pet in pets_filtrados
    )

    grafico_barras(
        dict(contagem_especies),
        "Distribuição por espécie",
        "Quantidade",
    )

with col_g2:
    st.subheader("📋 Classificações")

    contagem_classificacoes = Counter(
        pet["classificacao"] or "Sem avaliação"
        for pet in pets_filtrados
    )

    grafico_barras(
        dict(contagem_classificacoes),
        "Distribuição por classificação",
        "Quantidade",
    )

st.divider()

st.subheader("📈 Desempenho das avaliações")

pets_avaliados = [
    pet
    for pet in pets_filtrados
    if pet["classificacao"] != "Sem avaliação"
]

grafico_medias(pets_avaliados)

# ==========================================================
# DESTAQUES
# ==========================================================

st.divider()
st.subheader("🏆 Destaques")

if pets_avaliados:
    melhor = max(
        pets_avaliados,
        key=lambda item: item["media"],
    )
    menor = min(
        pets_avaliados,
        key=lambda item: item["media"],
    )

    col_d1, col_d2, col_d3 = st.columns(3)

    with col_d1:
        st.markdown(
            f"""
            <div class="painel-card">
                <h3>🏆 Melhor resultado</h3>
                <p><b>Pet:</b> {melhor["nome"]}</p>
                <p><b>Média:</b> {melhor["media"]:.2f}</p>
                <p><b>Classificação:</b> {melhor["classificacao"]}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_d2:
        st.markdown(
            f"""
            <div class="painel-card">
                <h3>🔎 Menor resultado</h3>
                <p><b>Pet:</b> {menor["nome"]}</p>
                <p><b>Média:</b> {menor["media"]:.2f}</p>
                <p><b>Classificação:</b> {menor["classificacao"]}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_d3:
        ultimo = max(
            pets_avaliados,
            key=lambda item: item["arquivo"].stat().st_mtime,
        )

        st.markdown(
            f"""
            <div class="painel-card">
                <h3>🕒 Última atualização</h3>
                <p><b>Pet:</b> {ultimo["nome"]}</p>
                <p><b>Data:</b> {ultimo["data_avaliacao"] or "-"}</p>
                <p><b>Veterinário:</b> {ultimo["veterinario"] or "-"}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
else:
    st.info(
        "Nenhum pet possui avaliação salva para exibir destaques."
    )

# ==========================================================
# TABELA
# ==========================================================

st.divider()
st.subheader("📄 Resumo dos pets")

linhas = []

for pet in pets_filtrados:
    linhas.append(
        {
            "Pet": pet["nome"],
            "Espécie": pet["especie"] or "-",
            "Raça": pet["raca"] or "-",
            "Sexo": pet["sexo"] or "-",
            "Peso (kg)": round(pet["peso"], 2),
            "Tutor": pet["tutor"] or "-",
            "Média": round(pet["media"], 2),
            "Classificação": pet["classificacao"],
            "Última avaliação": pet["data_avaliacao"] or "-",
        }
    )

df = pd.DataFrame(linhas)

st.dataframe(
    df,
    use_container_width=True,
    hide_index=True,
)

# ==========================================================
# AÇÕES
# ==========================================================

st.divider()

col_a1, col_a2, col_a3 = st.columns(3)

with col_a1:
    if st.button(
        "➕ Novo cadastro",
        use_container_width=True,
    ):
        st.switch_page("pages/Cadastro.py")

with col_a2:
    if st.button(
        "🩺 Nova avaliação",
        use_container_width=True,
    ):
        st.switch_page("pages/Avaliacao.py")

with col_a3:
    if st.button(
        "🐾 Abrir Pets",
        use_container_width=True,
    ):
        st.switch_page("pages/Pets.py")

# ==========================================================
# RODAPÉ
# ==========================================================

st.divider()
st.caption(
    "DinoPetAnimal PRO • Dashboard integrado às planilhas individuais"
)
