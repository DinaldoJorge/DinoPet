from pathlib import Path
from shutil import copy2
from datetime import date, datetime
import re

import streamlit as st
from openpyxl import load_workbook

# ==========================================================
# DINOPETANIMAL PRO
# PÁGINA: CADASTRO DE PET
# ==========================================================


# ==========================================================
# CAMINHOS DO PROJETO
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_PETS = BASE_DIR / "pets"
MODELO_PET = BASE_DIR / "modelo_pet.xlsx"
ARQUIVO_CSS = BASE_DIR / "styles.css"

PASTA_PETS.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CSS GLOBAL
# ==========================================================

if ARQUIVO_CSS.exists() and ARQUIVO_CSS.stat().st_size > 0:
    st.markdown(
        f"<style>{ARQUIVO_CSS.read_text(encoding='utf-8')}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    .cadastro-topo {
        background: linear-gradient(135deg, #fff7ed, #ffedd5);
        border: 1px solid #fdba74;
        border-radius: 18px;
        padding: 22px 26px;
        margin-bottom: 22px;
    }

    .cadastro-topo h1 {
        margin: 0;
        color: #1f2937;
        font-size: 2.2rem;
    }

    .cadastro-topo p {
        margin: 8px 0 0 0;
        color: #475569;
        font-size: 1.05rem;
    }

    .secao-titulo {
        font-size: 1.25rem;
        font-weight: 700;
        color: #7c2d12;
        margin-top: 12px;
        margin-bottom: 4px;
    }

    div[data-testid="stForm"] {
        border: 1px solid #fed7aa;
        border-radius: 18px;
        padding: 22px;
        background: #ffffff;
    }

    .stButton > button,
    .stFormSubmitButton > button {
        border-radius: 10px;
        min-height: 46px;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# FUNÇÕES AUXILIARES
# ==========================================================

def texto_limpo(valor) -> str:
    """Remove espaços extras e converte o valor para texto."""
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def nome_seguro_arquivo(nome: str) -> str:
    """
    Gera um nome seguro para o arquivo Excel.
    Mantém letras, números, espaços, hífen e sublinhado.
    """
    nome = texto_limpo(nome)
    nome = re.sub(r'[\\/:*?"<>|]', "", nome)
    nome = nome.rstrip(". ")
    return nome


def calcular_idade(data_nascimento: date) -> str:
    """Calcula a idade aproximada em anos e meses."""
    hoje = date.today()

    if data_nascimento > hoje:
        return ""

    anos = hoje.year - data_nascimento.year
    meses = hoje.month - data_nascimento.month

    if hoje.day < data_nascimento.day:
        meses -= 1

    if meses < 0:
        anos -= 1
        meses += 12

    if anos <= 0:
        return f"{meses} mês" if meses == 1 else f"{meses} meses"

    if meses == 0:
        return f"{anos} ano" if anos == 1 else f"{anos} anos"

    texto_anos = f"{anos} ano" if anos == 1 else f"{anos} anos"
    texto_meses = f"{meses} mês" if meses == 1 else f"{meses} meses"
    return f"{texto_anos} e {texto_meses}"


def caminho_pet(nome: str) -> Path:
    """Retorna o caminho do arquivo individual do pet."""
    return PASTA_PETS / f"{nome_seguro_arquivo(nome)}.xlsx"


def localizar_linha(ws, rotulo: str):
    """
    Localiza a linha de um campo pelo texto da coluna A.
    Isso mantém o cadastro funcionando mesmo se as linhas
    do modelo forem reorganizadas.
    """
    procurado = texto_limpo(rotulo).casefold()

    for linha in range(1, ws.max_row + 1):
        atual = texto_limpo(ws.cell(linha, 1).value).casefold()
        if atual == procurado:
            return linha

    return None


def escrever_campo(ws, rotulo: str, valor):
    """Escreve o valor na coluna B ao lado do rótulo informado."""
    linha = localizar_linha(ws, rotulo)

    if linha is None:
        raise KeyError(f"Campo não encontrado no modelo: {rotulo}")

    ws.cell(linha, 2).value = valor


def salvar_novo_pet(dados: dict) -> Path:
    """Cria a planilha individual do pet com base no modelo."""
    nome = nome_seguro_arquivo(dados["nome"])

    if len(nome) < 2:
        raise ValueError("Informe um nome válido para o pet.")

    if not MODELO_PET.exists():
        raise FileNotFoundError(
            f"O arquivo modelo_pet.xlsx não foi encontrado em:\n{MODELO_PET}"
        )

    destino = caminho_pet(nome)

    if destino.exists():
        raise FileExistsError(
            f"Já existe um pet chamado '{nome}'. "
            "Abra a página Pets para consultar ou editar o cadastro."
        )

    copy2(MODELO_PET, destino)

    try:
        wb = load_workbook(destino)

        if "Cadastro" not in wb.sheetnames:
            raise KeyError("A aba 'Cadastro' não existe no modelo_pet.xlsx.")

        ws = wb["Cadastro"]

        campos = {
            "Nome do Pet": nome,
            "Microchip": dados["microchip"],
            "Espécie": dados["especie"],
            "Raça": dados["raca"],
            "Sexo": dados["sexo"],
            "Castrado": dados["castrado"],
            "Data de Nascimento": dados["nascimento"].strftime("%d/%m/%Y"),
            "Idade": dados["idade"],
            "Cor": dados["cor"],
            "Peso (kg)": dados["peso"],
            "Tutor": dados["tutor"],
            "CPF": dados["cpf"],
            "Telefone": dados["telefone"],
            "E-mail": dados["email"],
            "Endereço": dados["endereco"],
            "Veterinário": dados["veterinario"],
            "CRMV": dados["crmv"],
            "Observações": dados["observacoes"],
        }

        for rotulo, valor in campos.items():
            escrever_campo(ws, rotulo, valor)

        # Registros internos, quando a aba existir.
        if "Oculto" in wb.sheetnames:
            ws_oculto = wb["Oculto"]
            agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            if ws_oculto.max_row >= 2:
                ws_oculto["A2"] = f"PET-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
            if ws_oculto.max_row >= 3:
                ws_oculto["A3"] = agora
            if ws_oculto.max_row >= 4:
                ws_oculto["A4"] = agora
            if ws_oculto.max_row >= 5:
                ws_oculto["A5"] = "5.0"

        wb.save(destino)
        wb.close()

    except Exception:
        # Evita deixar um arquivo incompleto em caso de falha.
        if destino.exists():
            destino.unlink()
        raise

    return destino


def limpar_formulario():
    """Reinicia os campos controlados pelo session_state."""
    chaves = [
        "cad_nome",
        "cad_microchip",
        "cad_raca",
        "cad_cor",
        "cad_peso",
        "cad_tutor",
        "cad_cpf",
        "cad_telefone",
        "cad_email",
        "cad_endereco",
        "cad_veterinario",
        "cad_crmv",
        "cad_observacoes",
    ]

    for chave in chaves:
        st.session_state.pop(chave, None)


# ==========================================================
# CABEÇALHO
# ==========================================================

st.markdown(
    """
    <div class="cadastro-topo">
        <h1>➕ Cadastro de Pet</h1>
        <p>Preencha os dados do animal e do tutor para criar uma planilha individual.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

if not MODELO_PET.exists():
    st.error(
        "❌ O arquivo **modelo_pet.xlsx** não foi encontrado na pasta principal "
        "do projeto. O cadastro não poderá ser salvo."
    )
    st.stop()

# ==========================================================
# FORMULÁRIO
# ==========================================================

with st.form("formulario_cadastro_pet", clear_on_submit=False):

    st.markdown(
        '<div class="secao-titulo">🐾 Dados do animal</div>',
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        nome = st.text_input(
            "Nome do pet *",
            key="cad_nome",
            placeholder="Ex.: Dino",
        )

        especie = st.selectbox(
            "Espécie *",
            [
                "Cão",
                "Gato",
                "Ave",
                "Coelho",
                "Roedor",
                "Réptil",
                "Outro",
            ],
        )

        sexo = st.selectbox(
            "Sexo",
            ["Macho", "Fêmea", "Não informado"],
        )

    with col2:
        raca = st.text_input(
            "Raça",
            key="cad_raca",
            placeholder="Ex.: Golden Retriever",
        )

        cor = st.text_input(
            "Cor",
            key="cad_cor",
            placeholder="Ex.: Dourado",
        )

        castrado = st.selectbox(
            "Castrado",
            ["Não", "Sim", "Não informado"],
        )

    with col3:
        nascimento = st.date_input(
            "Data de nascimento",
            value=date.today(),
            max_value=date.today(),
            format="DD/MM/YYYY",
        )

        idade_calculada = calcular_idade(nascimento)

        st.text_input(
            "Idade calculada",
            value=idade_calculada,
            disabled=True,
        )

        peso = st.number_input(
            "Peso (kg)",
            min_value=0.0,
            max_value=500.0,
            value=0.0,
            step=0.1,
            format="%.2f",
            key="cad_peso",
        )

    microchip = st.text_input(
        "Número do microchip",
        key="cad_microchip",
        placeholder="Deixe em branco quando não houver",
    )

    st.divider()

    st.markdown(
        '<div class="secao-titulo">👤 Dados do tutor</div>',
        unsafe_allow_html=True,
    )

    col4, col5 = st.columns(2)

    with col4:
        tutor = st.text_input(
            "Nome completo do tutor *",
            key="cad_tutor",
            placeholder="Nome do responsável",
        )

        cpf = st.text_input(
            "CPF",
            key="cad_cpf",
            placeholder="000.000.000-00",
        )

        telefone = st.text_input(
            "Telefone/WhatsApp",
            key="cad_telefone",
            placeholder="(00) 00000-0000",
        )

    with col5:
        email = st.text_input(
            "E-mail",
            key="cad_email",
            placeholder="tutor@email.com",
        )

        endereco = st.text_area(
            "Endereço",
            key="cad_endereco",
            placeholder="Rua, número, bairro, cidade e estado",
            height=118,
        )

    st.divider()

    st.markdown(
        '<div class="secao-titulo">🩺 Atendimento veterinário</div>',
        unsafe_allow_html=True,
    )

    col6, col7 = st.columns(2)

    with col6:
        veterinario = st.text_input(
            "Veterinário responsável",
            key="cad_veterinario",
            placeholder="Nome do médico-veterinário",
        )

    with col7:
        crmv = st.text_input(
            "CRMV",
            key="cad_crmv",
            placeholder="Ex.: CRMV-PB 0000",
        )

    observacoes = st.text_area(
        "Observações",
        key="cad_observacoes",
        placeholder="Alergias, medicamentos, cuidados especiais ou outras informações.",
        height=130,
    )

    st.caption("* Campos obrigatórios.")

    enviar = st.form_submit_button(
        "💾 Salvar novo pet",
        type="primary",
        use_container_width=True,
    )

# ==========================================================
# SALVAMENTO
# ==========================================================

if enviar:
    nome_final = texto_limpo(nome)
    tutor_final = texto_limpo(tutor)

    erros = []

    if len(nome_final) < 2:
        erros.append("Informe o nome do pet.")

    if not texto_limpo(especie):
        erros.append("Selecione a espécie.")

    if len(tutor_final) < 2:
        erros.append("Informe o nome do tutor.")

    if peso <= 0:
        erros.append("Informe um peso maior que zero.")

    if email and ("@" not in email or "." not in email.split("@")[-1]):
        erros.append("O e-mail informado parece inválido.")

    if erros:
        for erro in erros:
            st.error(f"❌ {erro}")
    else:
        dados_pet = {
            "nome": nome_final,
            "microchip": texto_limpo(microchip),
            "especie": texto_limpo(especie),
            "raca": texto_limpo(raca),
            "sexo": texto_limpo(sexo),
            "castrado": texto_limpo(castrado),
            "nascimento": nascimento,
            "idade": idade_calculada,
            "cor": texto_limpo(cor),
            "peso": float(peso),
            "tutor": tutor_final,
            "cpf": texto_limpo(cpf),
            "telefone": texto_limpo(telefone),
            "email": texto_limpo(email),
            "endereco": texto_limpo(endereco),
            "veterinario": texto_limpo(veterinario),
            "crmv": texto_limpo(crmv),
            "observacoes": texto_limpo(observacoes),
        }

        try:
            arquivo_criado = salvar_novo_pet(dados_pet)

            st.session_state["ultimo_pet_cadastrado"] = nome_final

            st.success(
                f"✅ Pet **{nome_final}** cadastrado com sucesso!"
            )
            st.info(
                f"📁 Arquivo criado: **pets/{arquivo_criado.name}**"
            )

            col_sucesso1, col_sucesso2 = st.columns(2)

            with col_sucesso1:
                if st.button(
                    "🐕 Abrir página Pets",
                    use_container_width=True,
                ):
                    st.switch_page("pages/Pets.py")

            with col_sucesso2:
                if st.button(
                    "➕ Cadastrar outro pet",
                    use_container_width=True,
                ):
                    limpar_formulario()
                    st.rerun()

        except FileExistsError as erro:
            st.warning(f"⚠️ {erro}")

        except Exception as erro:
            st.error(
                "❌ Não foi possível salvar o cadastro."
            )
            st.exception(erro)

# ==========================================================
# RODAPÉ
# ==========================================================

st.divider()

col_rodape1, col_rodape2 = st.columns(2)

with col_rodape1:
    st.caption(
        f"📂 Pasta de cadastros: {PASTA_PETS.name}/"
    )

with col_rodape2:
    st.caption(
        "DinoPetAnimal PRO • Cadastro individual em Excel"
    )