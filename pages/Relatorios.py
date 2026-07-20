from pathlib import Path
from io import BytesIO
from datetime import datetime
import textwrap

import streamlit as st
import matplotlib.pyplot as plt
from openpyxl import load_workbook

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image,
        PageBreak,
        KeepTogether,
    )

    REPORTLAB_DISPONIVEL = True
except ImportError:
    REPORTLAB_DISPONIVEL = False

# ==========================================================
# DINOPETANIMAL PRO
# PÁGINA: RELATÓRIOS
# ==========================================================


# ==========================================================
# CAMINHOS
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_PETS = BASE_DIR / "pets"
PASTA_RELATORIOS = BASE_DIR / "relatorios"
ARQUIVO_CSS = BASE_DIR / "styles.css"
ARQUIVO_LOGO = BASE_DIR / "animal.png"

PASTA_PETS.mkdir(parents=True, exist_ok=True)
PASTA_RELATORIOS.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CSS
# ==========================================================

if ARQUIVO_CSS.exists() and ARQUIVO_CSS.stat().st_size > 0:
    st.markdown(
        f"<style>{ARQUIVO_CSS.read_text(encoding='utf-8')}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    .relatorio-topo {
        background: linear-gradient(135deg, #fff7ed, #ffedd5);
        border: 1px solid #fdba74;
        border-radius: 18px;
        padding: 22px 26px;
        margin-bottom: 22px;
    }

    .relatorio-topo h1 {
        margin: 0;
        color: #1f2937;
        font-size: 2.2rem;
    }

    .relatorio-topo p {
        margin: 8px 0 0 0;
        color: #475569;
        font-size: 1.05rem;
    }

    .pet-card {
        background: #ffffff;
        border: 1px solid #fed7aa;
        border-radius: 16px;
        padding: 18px 20px;
        margin-bottom: 16px;
        box-shadow: 0 2px 10px rgba(0, 0, 0, .04);
    }

    .pet-card h3 {
        margin: 0 0 8px 0;
        color: #7c2d12;
    }

    .pet-card p {
        margin: 4px 0;
        color: #475569;
    }

    .stButton > button,
    .stDownloadButton > button {
        border-radius: 10px;
        min-height: 46px;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# FUNÇÕES DE LEITURA
# ==========================================================

def texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def listar_arquivos_pets():
    return sorted(
        [
            arquivo
            for arquivo in PASTA_PETS.glob("*.xlsx")
            if not arquivo.name.startswith("~$")
        ],
        key=lambda arquivo: arquivo.stem.casefold(),
    )


def localizar_linha(ws, rotulo: str):
    procurado = texto_limpo(rotulo).casefold()

    for linha in range(1, ws.max_row + 1):
        atual = texto_limpo(ws.cell(linha, 1).value).casefold()
        if atual == procurado:
            return linha

    return None


def ler_campo(ws, rotulo: str, coluna=2, padrao=""):
    linha = localizar_linha(ws, rotulo)

    if linha is None:
        return padrao

    valor = ws.cell(linha, coluna).value
    return padrao if valor is None else valor


def converter_numero(valor, padrao=0.0):
    try:
        return float(valor)
    except (TypeError, ValueError):
        return padrao


def carregar_dados_pet(arquivo: Path) -> dict:
    wb = load_workbook(arquivo, data_only=True)

    if "Cadastro" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba Cadastro não foi encontrada.")

    ws_cadastro = wb["Cadastro"]

    cadastro = {
        "nome": texto_limpo(
            ler_campo(ws_cadastro, "Nome do Pet", 2, arquivo.stem)
        ),
        "microchip": texto_limpo(
            ler_campo(ws_cadastro, "Microchip")
        ),
        "especie": texto_limpo(
            ler_campo(ws_cadastro, "Espécie")
        ),
        "raca": texto_limpo(
            ler_campo(ws_cadastro, "Raça")
        ),
        "sexo": texto_limpo(
            ler_campo(ws_cadastro, "Sexo")
        ),
        "castrado": texto_limpo(
            ler_campo(ws_cadastro, "Castrado")
        ),
        "nascimento": texto_limpo(
            ler_campo(ws_cadastro, "Data de Nascimento")
        ),
        "idade": texto_limpo(
            ler_campo(ws_cadastro, "Idade")
        ),
        "cor": texto_limpo(
            ler_campo(ws_cadastro, "Cor")
        ),
        "peso": converter_numero(
            ler_campo(ws_cadastro, "Peso (kg)", 2, 0)
        ),
        "tutor": texto_limpo(
            ler_campo(ws_cadastro, "Tutor")
        ),
        "cpf": texto_limpo(
            ler_campo(ws_cadastro, "CPF")
        ),
        "telefone": texto_limpo(
            ler_campo(ws_cadastro, "Telefone")
        ),
        "email": texto_limpo(
            ler_campo(ws_cadastro, "E-mail")
        ),
        "endereco": texto_limpo(
            ler_campo(ws_cadastro, "Endereço")
        ),
        "veterinario": texto_limpo(
            ler_campo(ws_cadastro, "Veterinário")
        ),
        "crmv": texto_limpo(
            ler_campo(ws_cadastro, "CRMV")
        ),
        "observacoes": texto_limpo(
            ler_campo(ws_cadastro, "Observações")
        ),
    }

    avaliacoes = []
    observacao_geral = ""

    if "Avaliacao" in wb.sheetnames:
        ws_avaliacao = wb["Avaliacao"]

        for linha in range(2, ws_avaliacao.max_row + 1):
            dominio = texto_limpo(ws_avaliacao.cell(linha, 1).value)

            if not dominio:
                continue

            if dominio.casefold() == "observação geral".casefold():
                observacao_geral = texto_limpo(
                    ws_avaliacao.cell(linha, 2).value
                )
                continue

            nota = converter_numero(
                ws_avaliacao.cell(linha, 2).value,
                0,
            )
            peso = converter_numero(
                ws_avaliacao.cell(linha, 3).value,
                1,
            )
            resultado = converter_numero(
                ws_avaliacao.cell(linha, 4).value,
                nota * peso,
            )

            avaliacoes.append(
                {
                    "dominio": dominio,
                    "nota": nota,
                    "peso": peso,
                    "resultado": resultado,
                }
            )

    resultados = {
        "score": 0.0,
        "media": 0.0,
        "maior": 0.0,
        "menor": 0.0,
        "classificacao": "Sem avaliação",
        "data": "",
        "responsavel": "",
    }

    if "Resultados" in wb.sheetnames:
        ws_resultados = wb["Resultados"]

        resultados = {
            "score": converter_numero(
                ler_campo(ws_resultados, "Score Geral", 2, 0)
            ),
            "media": converter_numero(
                ler_campo(ws_resultados, "Média", 2, 0)
            ),
            "maior": converter_numero(
                ler_campo(ws_resultados, "Maior Nota", 2, 0)
            ),
            "menor": converter_numero(
                ler_campo(ws_resultados, "Menor Nota", 2, 0)
            ),
            "classificacao": texto_limpo(
                ler_campo(
                    ws_resultados,
                    "Classificação",
                    2,
                    "Sem avaliação",
                )
            ) or "Sem avaliação",
            "data": texto_limpo(
                ler_campo(ws_resultados, "Data da Avaliação")
            ),
            "responsavel": texto_limpo(
                ler_campo(ws_resultados, "Responsável")
            ),
        }

    wb.close()

    return {
        "arquivo": arquivo,
        "cadastro": cadastro,
        "avaliacoes": avaliacoes,
        "observacao_geral": observacao_geral,
        "resultados": resultados,
    }


# ==========================================================
# GRÁFICO
# ==========================================================

def criar_grafico_avaliacao(avaliacoes: list[dict]) -> BytesIO | None:
    avaliacoes_validas = [
        item for item in avaliacoes if item["nota"] > 0
    ]

    if not avaliacoes_validas:
        return None

    dominios = [item["dominio"] for item in avaliacoes_validas]
    notas = [item["nota"] for item in avaliacoes_validas]

    altura = max(5.0, len(dominios) * 0.38)

    fig, ax = plt.subplots(figsize=(9, altura))
    ax.barh(dominios, notas)
    ax.set_xlim(0, 10)
    ax.set_xlabel("Nota")
    ax.set_title("Resultado por domínio")
    ax.grid(axis="x", alpha=0.25)
    ax.invert_yaxis()

    for indice, nota in enumerate(notas):
        ax.text(
            nota + 0.10,
            indice,
            f"{nota:.1f}",
            va="center",
            fontsize=9,
        )

    fig.tight_layout()

    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=160,
        bbox_inches="tight",
    )
    plt.close(fig)

    buffer.seek(0)
    return buffer


# ==========================================================
# GERAÇÃO DO PDF
# ==========================================================

def quebrar_texto(texto: str, limite=80) -> str:
    texto = texto_limpo(texto)
    if not texto:
        return "-"

    return "<br/>".join(
        textwrap.wrap(
            texto,
            width=limite,
            break_long_words=False,
            replace_whitespace=False,
        )
    )


def gerar_pdf(dados: dict) -> bytes:
    if not REPORTLAB_DISPONIVEL:
        raise RuntimeError(
            "A biblioteca reportlab não está instalada. "
            "Execute: pip install reportlab"
        )

    cadastro = dados["cadastro"]
    resultados = dados["resultados"]
    avaliacoes = dados["avaliacoes"]

    buffer_pdf = BytesIO()

    documento = SimpleDocTemplate(
        buffer_pdf,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=f"Relatório Veterinário - {cadastro['nome']}",
        author="DinoPetAnimal PRO",
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloDinoPet",
        parent=estilos["Title"],
        fontSize=19,
        leading=23,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#7C2D12"),
        spaceAfter=8,
    )

    estilo_subtitulo = ParagraphStyle(
        "SubtituloDinoPet",
        parent=estilos["Heading2"],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#9A3412"),
        spaceBefore=8,
        spaceAfter=8,
    )

    estilo_texto = ParagraphStyle(
        "TextoDinoPet",
        parent=estilos["BodyText"],
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
    )

    estilo_central = ParagraphStyle(
        "CentralDinoPet",
        parent=estilo_texto,
        alignment=TA_CENTER,
    )

    elementos = []

    if ARQUIVO_LOGO.exists():
        try:
            logo = Image(
                str(ARQUIVO_LOGO),
                width=17.5 * cm,
                height=4.4 * cm,
            )
            logo.hAlign = "CENTER"
            elementos.append(logo)
            elementos.append(Spacer(1, 0.20 * cm))
        except Exception:
            pass

    elementos.append(
        Paragraph(
            "RELATÓRIO VETERINÁRIO INDIVIDUAL",
            estilo_titulo,
        )
    )

    elementos.append(
        Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            estilo_central,
        )
    )

    elementos.append(Spacer(1, 0.35 * cm))

    # Dados do animal
    elementos.append(
        Paragraph("1. Identificação do animal", estilo_subtitulo)
    )

    tabela_animal = [
        [
            Paragraph("<b>Nome do pet</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["nome"]), estilo_texto),
            Paragraph("<b>Espécie</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["especie"]), estilo_texto),
        ],
        [
            Paragraph("<b>Raça</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["raca"]), estilo_texto),
            Paragraph("<b>Sexo</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["sexo"]), estilo_texto),
        ],
        [
            Paragraph("<b>Nascimento</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["nascimento"]), estilo_texto),
            Paragraph("<b>Idade</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["idade"]), estilo_texto),
        ],
        [
            Paragraph("<b>Peso</b>", estilo_texto),
            Paragraph(f"{cadastro['peso']:.2f} kg", estilo_texto),
            Paragraph("<b>Castrado</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["castrado"]), estilo_texto),
        ],
        [
            Paragraph("<b>Cor</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["cor"]), estilo_texto),
            Paragraph("<b>Microchip</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["microchip"]), estilo_texto),
        ],
    ]

    tabela = Table(
        tabela_animal,
        colWidths=[3.0 * cm, 5.2 * cm, 2.5 * cm, 6.0 * cm],
    )

    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDBA74")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabela)

    # Tutor
    elementos.append(
        Paragraph("2. Dados do tutor", estilo_subtitulo)
    )

    tabela_tutor = [
        [
            Paragraph("<b>Tutor</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["tutor"]), estilo_texto),
            Paragraph("<b>Telefone</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["telefone"]), estilo_texto),
        ],
        [
            Paragraph("<b>CPF</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["cpf"]), estilo_texto),
            Paragraph("<b>E-mail</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["email"]), estilo_texto),
        ],
        [
            Paragraph("<b>Endereço</b>", estilo_texto),
            Paragraph(
                quebrar_texto(cadastro["endereco"], 110),
                estilo_texto,
            ),
            "",
            "",
        ],
    ]

    tabela = Table(
        tabela_tutor,
        colWidths=[2.5 * cm, 5.7 * cm, 2.5 * cm, 6.0 * cm],
    )

    tabela.setStyle(
        TableStyle(
            [
                ("SPAN", (1, 2), (3, 2)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDBA74")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabela)

    # Atendimento
    elementos.append(
        Paragraph("3. Atendimento veterinário", estilo_subtitulo)
    )

    tabela_veterinario = [
        [
            Paragraph("<b>Veterinário</b>", estilo_texto),
            Paragraph(
                quebrar_texto(cadastro["veterinario"]),
                estilo_texto,
            ),
            Paragraph("<b>CRMV</b>", estilo_texto),
            Paragraph(quebrar_texto(cadastro["crmv"]), estilo_texto),
        ],
        [
            Paragraph("<b>Responsável pela avaliação</b>", estilo_texto),
            Paragraph(
                quebrar_texto(resultados["responsavel"]),
                estilo_texto,
            ),
            Paragraph("<b>Data da avaliação</b>", estilo_texto),
            Paragraph(quebrar_texto(resultados["data"]), estilo_texto),
        ],
    ]

    tabela = Table(
        tabela_veterinario,
        colWidths=[3.8 * cm, 4.5 * cm, 3.4 * cm, 5.0 * cm],
    )

    tabela.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDBA74")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabela)

    # Resultados
    elementos.append(
        Paragraph("4. Resultado geral", estilo_subtitulo)
    )

    tabela_resultado = [
        [
            Paragraph("<b>Score geral</b>", estilo_central),
            Paragraph("<b>Média</b>", estilo_central),
            Paragraph("<b>Maior nota</b>", estilo_central),
            Paragraph("<b>Menor nota</b>", estilo_central),
            Paragraph("<b>Classificação</b>", estilo_central),
        ],
        [
            Paragraph(f"{resultados['score']:.2f}", estilo_central),
            Paragraph(f"{resultados['media']:.2f}", estilo_central),
            Paragraph(f"{resultados['maior']:.2f}", estilo_central),
            Paragraph(f"{resultados['menor']:.2f}", estilo_central),
            Paragraph(
                quebrar_texto(resultados["classificacao"]),
                estilo_central,
            ),
        ],
    ]

    tabela = Table(
        tabela_resultado,
        colWidths=[3.1 * cm, 3.1 * cm, 3.1 * cm, 3.1 * cm, 4.3 * cm],
    )

    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9A3412")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FFEDD5")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#FDBA74")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )

    elementos.append(tabela)

    # Avaliação detalhada
    elementos.append(
        Paragraph("5. Avaliação por domínio", estilo_subtitulo)
    )

    linhas_avaliacao = [
        [
            Paragraph("<b>Domínio</b>", estilo_central),
            Paragraph("<b>Nota</b>", estilo_central),
            Paragraph("<b>Peso</b>", estilo_central),
            Paragraph("<b>Resultado</b>", estilo_central),
        ]
    ]

    for item in avaliacoes:
        linhas_avaliacao.append(
            [
                Paragraph(
                    quebrar_texto(item["dominio"], 45),
                    estilo_texto,
                ),
                Paragraph(f"{item['nota']:.1f}", estilo_central),
                Paragraph(f"{item['peso']:.1f}", estilo_central),
                Paragraph(f"{item['resultado']:.2f}", estilo_central),
            ]
        )

    tabela = Table(
        linhas_avaliacao,
        colWidths=[9.5 * cm, 2.3 * cm, 2.3 * cm, 2.8 * cm],
        repeatRows=1,
    )

    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9A3412")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#FDBA74")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                    colors.white,
                    colors.HexColor("#FFF7ED"),
                ]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elementos.append(tabela)

    # Gráfico
    grafico = criar_grafico_avaliacao(avaliacoes)

    if grafico is not None:
        elementos.append(Spacer(1, 0.35 * cm))
        elementos.append(
            Paragraph("6. Gráfico da avaliação", estilo_subtitulo)
        )

        imagem_grafico = Image(
            grafico,
            width=17.0 * cm,
            height=9.0 * cm,
        )
        imagem_grafico.hAlign = "CENTER"
        elementos.append(imagem_grafico)

    # Observações
    elementos.append(
        Paragraph("7. Observações", estilo_subtitulo)
    )

    observacoes = (
        dados["observacao_geral"]
        or cadastro["observacoes"]
        or "Nenhuma observação registrada."
    )

    tabela_observacoes = Table(
        [[Paragraph(quebrar_texto(observacoes, 120), estilo_texto)]],
        colWidths=[16.9 * cm],
    )

    tabela_observacoes.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#FDBA74")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )

    elementos.append(tabela_observacoes)
    elementos.append(Spacer(1, 1.0 * cm))

    assinatura = Table(
        [
            [
                "",
                "",
            ],
            [
                Paragraph(
                    "_________________________________________<br/>"
                    "Responsável pela avaliação",
                    estilo_central,
                ),
                Paragraph(
                    "_________________________________________<br/>"
                    "Tutor ou responsável",
                    estilo_central,
                ),
            ],
        ],
        colWidths=[8.2 * cm, 8.2 * cm],
    )

    assinatura.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("TOPPADDING", (0, 0), (-1, -1), 15),
            ]
        )
    )

    elementos.append(assinatura)
    elementos.append(Spacer(1, 0.5 * cm))

    elementos.append(
        Paragraph(
            "Documento gerado pelo DinoPetAnimal PRO.",
            estilo_central,
        )
    )

    documento.build(elementos)

    buffer_pdf.seek(0)
    return buffer_pdf.getvalue()


# ==========================================================
# INTERFACE
# ==========================================================

st.markdown(
    """
    <div class="relatorio-topo">
        <h1>📄 Relatórios</h1>
        <p>Visualize os resultados e gere o relatório veterinário individual em PDF.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

arquivos = listar_arquivos_pets()

if not arquivos:
    st.info(
        "Nenhum pet foi cadastrado. Cadastre um animal antes de gerar relatórios."
    )

    if st.button(
        "➕ Ir para Cadastro",
        type="primary",
        use_container_width=True,
    ):
        st.switch_page("pages/Cadastro.py")

    st.stop()

pets = []

for arquivo in arquivos:
    try:
        dados = carregar_dados_pet(arquivo)
        pets.append(dados)
    except Exception as erro:
        st.warning(
            f"Não foi possível abrir {arquivo.name}: {erro}"
        )

if not pets:
    st.error("Nenhuma planilha válida foi encontrada.")
    st.stop()

nomes = [item["cadastro"]["nome"] for item in pets]

indice_inicial = 0
pet_memoria = st.session_state.get("pet_selecionado")

if pet_memoria:
    for indice, nome in enumerate(nomes):
        if nome.casefold() == str(pet_memoria).casefold():
            indice_inicial = indice
            break

nome_selecionado = st.selectbox(
    "🐾 Selecione o pet",
    nomes,
    index=indice_inicial,
)

dados = pets[nomes.index(nome_selecionado)]
cadastro = dados["cadastro"]
resultados = dados["resultados"]
avaliacoes = dados["avaliacoes"]

st.session_state["pet_selecionado"] = cadastro["nome"]

st.markdown(
    f"""
    <div class="pet-card">
        <h3>🐾 {cadastro["nome"]}</h3>
        <p><b>Espécie:</b> {cadastro["especie"] or "-"}</p>
        <p><b>Raça:</b> {cadastro["raca"] or "-"}</p>
        <p><b>Tutor:</b> {cadastro["tutor"] or "-"}</p>
        <p><b>Veterinário:</b> {cadastro["veterinario"] or "-"}</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# INDICADORES
# ==========================================================

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.metric("📊 Score geral", f"{resultados['score']:.2f}")

with col2:
    st.metric("📈 Média", f"{resultados['media']:.2f}")

with col3:
    st.metric("⬆️ Maior nota", f"{resultados['maior']:.2f}")

with col4:
    st.metric(
        "🏷️ Classificação",
        resultados["classificacao"],
    )

if resultados["data"]:
    st.caption(
        f"Última avaliação: {resultados['data']} "
        f"• Responsável: {resultados['responsavel'] or '-'}"
    )

st.divider()

# ==========================================================
# PRÉVIA
# ==========================================================

st.subheader("📋 Prévia dos dados")

col_p1, col_p2 = st.columns(2)

with col_p1:
    st.markdown("#### 🐾 Animal")
    st.write(f"**Nome:** {cadastro['nome'] or '-'}")
    st.write(f"**Espécie:** {cadastro['especie'] or '-'}")
    st.write(f"**Raça:** {cadastro['raca'] or '-'}")
    st.write(f"**Sexo:** {cadastro['sexo'] or '-'}")
    st.write(f"**Idade:** {cadastro['idade'] or '-'}")
    st.write(f"**Peso:** {cadastro['peso']:.2f} kg")

with col_p2:
    st.markdown("#### 👤 Tutor e atendimento")
    st.write(f"**Tutor:** {cadastro['tutor'] or '-'}")
    st.write(f"**Telefone:** {cadastro['telefone'] or '-'}")
    st.write(f"**E-mail:** {cadastro['email'] or '-'}")
    st.write(f"**Veterinário:** {cadastro['veterinario'] or '-'}")
    st.write(f"**CRMV:** {cadastro['crmv'] or '-'}")
    st.write(
        f"**Responsável pela avaliação:** "
        f"{resultados['responsavel'] or '-'}"
    )

st.divider()

st.subheader("🩺 Avaliação por domínio")

linhas_tabela = []

for item in avaliacoes:
    linhas_tabela.append(
        {
            "Domínio": item["dominio"],
            "Nota": round(item["nota"], 1),
            "Peso": round(item["peso"], 1),
            "Resultado": round(item["resultado"], 2),
        }
    )

if linhas_tabela:
    st.dataframe(
        linhas_tabela,
        use_container_width=True,
        hide_index=True,
    )
else:
    st.info("O pet ainda não possui domínios de avaliação.")

grafico_preview = criar_grafico_avaliacao(avaliacoes)

if grafico_preview is not None:
    st.subheader("📈 Gráfico")
    imagem = plt.imread(grafico_preview)
    st.image(imagem, use_container_width=True)

st.divider()

# ==========================================================
# GERAÇÃO E DOWNLOAD
# ==========================================================

st.subheader("⬇️ Gerar documento")

if resultados["classificacao"] == "Sem avaliação":
    st.warning(
        "Este pet ainda não possui uma avaliação salva. "
        "O relatório poderá ser gerado, mas os resultados estarão zerados."
    )

if not REPORTLAB_DISPONIVEL:
    st.error(
        "A biblioteca **reportlab** não está instalada. "
        "Abra o terminal e execute: `pip install reportlab`"
    )
else:
    try:
        pdf = gerar_pdf(dados)

        nome_seguro = "".join(
            caractere
            for caractere in cadastro["nome"]
            if caractere.isalnum() or caractere in (" ", "-", "_")
        ).strip().replace(" ", "_")

        nome_pdf = (
            f"Relatorio_{nome_seguro}_"
            f"{datetime.now().strftime('%Y%m%d')}.pdf"
        )

        col_d1, col_d2 = st.columns(2)

        with col_d1:
            st.download_button(
                "📄 Baixar relatório em PDF",
                data=pdf,
                file_name=nome_pdf,
                mime="application/pdf",
                type="primary",
                use_container_width=True,
            )

        with col_d2:
            if st.button(
                "💾 Salvar cópia na pasta relatorios",
                use_container_width=True,
            ):
                destino = PASTA_RELATORIOS / nome_pdf
                destino.write_bytes(pdf)

                st.success(
                    f"✅ Relatório salvo em relatorios/{destino.name}"
                )

    except Exception as erro:
        st.error("❌ Não foi possível gerar o relatório.")
        st.exception(erro)

# ==========================================================
# AÇÕES
# ==========================================================

st.divider()

col_a1, col_a2, col_a3 = st.columns(3)

with col_a1:
    if st.button(
        "🩺 Abrir Avaliação",
        use_container_width=True,
    ):
        st.switch_page("pages/Avaliacao.py")

with col_a2:
    if st.button(
        "📊 Abrir Dashboard",
        use_container_width=True,
    ):
        st.switch_page("pages/Dashboard.py")

with col_a3:
    if st.button(
        "🐾 Abrir Pets",
        use_container_width=True,
    ):
        st.switch_page("pages/Pets.py")

# ==========================================================
# RODAPÉ
# ==========================================================

st.divider()
st.caption(
    "DinoPetAnimal PRO • Relatório veterinário individual em PDF"
)
