from pathlib import Path
from datetime import datetime
import math

import streamlit as st
from openpyxl import load_workbook

# ==========================================================
# DINOPETANIMAL PRO
# PÁGINA: AVALIAÇÃO
# ==========================================================


# ==========================================================
# CAMINHOS
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent.parent
PASTA_PETS = BASE_DIR / "pets"
ARQUIVO_CSS = BASE_DIR / "styles.css"

PASTA_PETS.mkdir(parents=True, exist_ok=True)

# ==========================================================
# CSS
# ==========================================================

if ARQUIVO_CSS.exists() and ARQUIVO_CSS.stat().st_size > 0:
    st.markdown(
        f"<style>{ARQUIVO_CSS.read_text(encoding='utf-8')}</style>",
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
    .avaliacao-topo {
        background: linear-gradient(135deg, #fff7ed, #ffedd5);
        border: 1px solid #fdba74;
        border-radius: 18px;
        padding: 22px 26px;
        margin-bottom: 22px;
    }

    .avaliacao-topo h1 {
        margin: 0;
        color: #1f2937;
        font-size: 2.2rem;
    }

    .avaliacao-topo p {
        margin: 8px 0 0 0;
        color: #475569;
        font-size: 1.05rem;
    }

    .pet-resumo {
        background: #ffffff;
        border: 1px solid #fed7aa;
        border-radius: 16px;
        padding: 16px 20px;
        margin-bottom: 18px;
    }

    .pet-resumo h3 {
        margin: 0 0 8px 0;
        color: #7c2d12;
    }

    .pet-resumo p {
        margin: 4px 0;
        color: #475569;
    }

    div[data-testid="stForm"] {
        border: 1px solid #fed7aa;
        border-radius: 18px;
        padding: 22px;
        background: #ffffff;
    }

    .stButton > button,
    .stFormSubmitButton > button {
        border-radius: 10px;
        min-height: 46px;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================================
# FUNÇÕES
# ==========================================================

def texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def listar_pets():
    return sorted(
        [
            arquivo
            for arquivo in PASTA_PETS.glob("*.xlsx")
            if not arquivo.name.startswith("~$")
        ],
        key=lambda arquivo: arquivo.stem.casefold(),
    )


def localizar_linha(ws, rotulo: str):
    procurado = texto_limpo(rotulo).casefold()

    for linha in range(1, ws.max_row + 1):
        atual = texto_limpo(ws.cell(linha, 1).value).casefold()
        if atual == procurado:
            return linha

    return None


def ler_campo(ws, rotulo: str, coluna=2, padrao=""):
    linha = localizar_linha(ws, rotulo)
    if linha is None:
        return padrao

    valor = ws.cell(linha, coluna).value
    return padrao if valor is None else valor


def carregar_resumo_pet(arquivo: Path) -> dict:
    wb = load_workbook(arquivo, data_only=False)

    if "Cadastro" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba Cadastro não foi encontrada.")

    ws = wb["Cadastro"]

    dados = {
        "nome": texto_limpo(ler_campo(ws, "Nome do Pet", 2, arquivo.stem)),
        "especie": texto_limpo(ler_campo(ws, "Espécie")),
        "raca": texto_limpo(ler_campo(ws, "Raça")),
        "tutor": texto_limpo(ler_campo(ws, "Tutor")),
        "veterinario": texto_limpo(ler_campo(ws, "Veterinário")),
    }

    wb.close()
    return dados


def carregar_avaliacao(arquivo: Path):
    wb = load_workbook(arquivo, data_only=False)

    if "Avaliacao" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba Avaliacao não foi encontrada.")

    ws = wb["Avaliacao"]

    dominios = []
    observacao = ""

    for linha in range(2, ws.max_row + 1):
        dominio = texto_limpo(ws.cell(linha, 1).value)

        if not dominio:
            continue

        if dominio.casefold() == "observação geral".casefold():
            observacao = texto_limpo(ws.cell(linha, 2).value)
            continue

        nota = ws.cell(linha, 2).value
        peso = ws.cell(linha, 3).value

        try:
            nota = float(nota)
        except (TypeError, ValueError):
            nota = 0.0

        try:
            peso = float(peso)
        except (TypeError, ValueError):
            peso = 1.0

        dominios.append(
            {
                "linha": linha,
                "dominio": dominio,
                "nota": max(0.0, min(10.0, nota)),
                "peso": max(0.1, min(10.0, peso)),
            }
        )

    resultados = {
        "score": "",
        "media": "",
        "maior": "",
        "menor": "",
        "classificacao": "",
        "data": "",
        "responsavel": "",
    }

    if "Resultados" in wb.sheetnames:
        ws_resultados = wb["Resultados"]
        resultados = {
            "score": ler_campo(ws_resultados, "Score Geral"),
            "media": ler_campo(ws_resultados, "Média"),
            "maior": ler_campo(ws_resultados, "Maior Nota"),
            "menor": ler_campo(ws_resultados, "Menor Nota"),
            "classificacao": texto_limpo(
                ler_campo(ws_resultados, "Classificação")
            ),
            "data": texto_limpo(
                ler_campo(ws_resultados, "Data da Avaliação")
            ),
            "responsavel": texto_limpo(
                ler_campo(ws_resultados, "Responsável")
            ),
        }

    wb.close()
    return dominios, observacao, resultados


def classificar(media: float) -> str:
    if media >= 9:
        return "Excelente"
    if media >= 7:
        return "Bom"
    if media >= 5:
        return "Regular"
    if media >= 3:
        return "Ruim"
    return "Crítico"


def calcular_resultados(avaliacoes: list[dict]) -> dict:
    if not avaliacoes:
        return {
            "score": 0.0,
            "media": 0.0,
            "maior": 0.0,
            "menor": 0.0,
            "classificacao": "Sem avaliação",
        }

    soma_pesos = sum(item["peso"] for item in avaliacoes)

    if soma_pesos <= 0:
        media = 0.0
    else:
        media = sum(
            item["nota"] * item["peso"] for item in avaliacoes
        ) / soma_pesos

    notas = [item["nota"] for item in avaliacoes]

    return {
        "score": round(media * 10, 2),
        "media": round(media, 2),
        "maior": round(max(notas), 2),
        "menor": round(min(notas), 2),
        "classificacao": classificar(media),
    }


def escrever_resultado(ws, rotulo: str, valor):
    linha = localizar_linha(ws, rotulo)
    if linha is None:
        raise KeyError(f"Campo não encontrado em Resultados: {rotulo}")
    ws.cell(linha, 2).value = valor


def salvar_avaliacao(
    arquivo: Path,
    avaliacoes: list[dict],
    observacao: str,
    responsavel: str,
):
    resultados = calcular_resultados(avaliacoes)

    wb = load_workbook(arquivo)

    if "Avaliacao" not in wb.sheetnames:
        wb.close()
        raise KeyError("A aba Avaliacao não existe na planilha.")

    ws = wb["Avaliacao"]

    for item in avaliacoes:
        linha = item["linha"]
        nota = float(item["nota"])
        peso = float(item["peso"])

        ws.cell(linha, 2).value = nota
        ws.cell(linha, 3).value = peso
        ws.cell(linha, 4).value = round(nota * peso, 2)

    linha_observacao = localizar_linha(ws, "Observação Geral")
    if linha_observacao:
        ws.cell(linha_observacao, 2).value = texto_limpo(observacao)

    if "Resultados" not in wb.sheetnames:
        wb.create_sheet("Resultados")

    ws_resultados = wb["Resultados"]

    escrever_resultado(
        ws_resultados,
        "Score Geral",
        resultados["score"],
    )
    escrever_resultado(
        ws_resultados,
        "Média",
        resultados["media"],
    )
    escrever_resultado(
        ws_resultados,
        "Maior Nota",
        resultados["maior"],
    )
    escrever_resultado(
        ws_resultados,
        "Menor Nota",
        resultados["menor"],
    )
    escrever_resultado(
        ws_resultados,
        "Classificação",
        resultados["classificacao"],
    )
    escrever_resultado(
        ws_resultados,
        "Data da Avaliação",
        datetime.now().strftime("%d/%m/%Y %H:%M"),
    )
    escrever_resultado(
        ws_resultados,
        "Responsável",
        texto_limpo(responsavel),
    )

    if "Config" in wb.sheetnames:
        ws_config = wb["Config"]
        configuracoes = {
            "Excelente": 9,
            "Bom": 7,
            "Regular": 5,
            "Ruim": 3,
            "Crítico": 0,
            "Nota Máxima": 10,
            "Peso Padrão": 1,
            "Versão": "1.0",
            "Modelo": "DinoPetAnimal PRO",
        }

        for rotulo, valor in configuracoes.items():
            linha = localizar_linha(ws_config, rotulo)
            if linha and ws_config.cell(linha, 2).value is None:
                ws_config.cell(linha, 2).value = valor

    if "Oculto" in wb.sheetnames:
        ws_oculto = wb["Oculto"]
        linha_alteracao = localizar_linha(
            ws_oculto,
            "Última Alteração",
        )
        if linha_alteracao:
            ws_oculto.cell(
                linha_alteracao,
                2,
            ).value = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    wb.save(arquivo)
    wb.close()

    return resultados


# ==========================================================
# CABEÇALHO
# ==========================================================

st.markdown(
    """
    <div class="avaliacao-topo">
        <h1>🩺 Avaliação do Pet</h1>
        <p>Registre as notas dos domínios e acompanhe o resultado geral do animal.</p>
    </div>
    """,
    unsafe_allow_html=True,
)

arquivos = listar_pets()

if not arquivos:
    st.info(
        "Nenhum pet foi cadastrado. Cadastre um animal antes de realizar a avaliação."
    )

    if st.button(
        "➕ Ir para Cadastro",
        type="primary",
        use_container_width=True,
    ):
        st.switch_page("pages/Cadastro.py")

    st.stop()

# ==========================================================
# ESCOLHA DO PET
# ==========================================================

resumos = []

for arquivo in arquivos:
    try:
        resumo = carregar_resumo_pet(arquivo)
        resumo["arquivo"] = arquivo
        resumos.append(resumo)
    except Exception as erro:
        st.warning(f"Não foi possível abrir {arquivo.name}: {erro}")

if not resumos:
    st.error("Nenhuma planilha válida foi encontrada na pasta pets.")
    st.stop()

nomes = [item["nome"] for item in resumos]

indice_inicial = 0
pet_memoria = st.session_state.get("pet_selecionado")

if pet_memoria:
    for indice, nome in enumerate(nomes):
        if nome.casefold() == str(pet_memoria).casefold():
            indice_inicial = indice
            break

nome_selecionado = st.selectbox(
    "🐾 Selecione o pet",
    nomes,
    index=indice_inicial,
)

pet = resumos[nomes.index(nome_selecionado)]
arquivo_pet = pet["arquivo"]

st.session_state["pet_selecionado"] = pet["nome"]

st.markdown(
    f"""
    <div class="pet-resumo">
        <h3>🐾 {pet["nome"]}</h3>
        <p><b>Espécie:</b> {pet["especie"] or "-"}</p>
        <p><b>Raça:</b> {pet["raca"] or "-"}</p>
        <p><b>Tutor:</b> {pet["tutor"] or "-"}</p>
        <p><b>Veterinário:</b> {pet["veterinario"] or "-"}</p>
    </div>
    """,
    unsafe_allow_html=True,
)

try:
    dominios, observacao_atual, resultados_atuais = (
        carregar_avaliacao(arquivo_pet)
    )
except Exception as erro:
    st.error("Não foi possível carregar a avaliação deste pet.")
    st.exception(erro)
    st.stop()

# ==========================================================
# RESULTADO ANTERIOR
# ==========================================================

if resultados_atuais["classificacao"]:
    st.subheader("📊 Último resultado salvo")

    col_r1, col_r2, col_r3, col_r4 = st.columns(4)

    with col_r1:
        st.metric(
            "Score geral",
            resultados_atuais["score"] or 0,
        )

    with col_r2:
        st.metric(
            "Média",
            resultados_atuais["media"] or 0,
        )

    with col_r3:
        st.metric(
            "Maior nota",
            resultados_atuais["maior"] or 0,
        )

    with col_r4:
        st.metric(
            "Classificação",
            resultados_atuais["classificacao"],
        )

    if resultados_atuais["data"]:
        st.caption(
            f"Última avaliação: {resultados_atuais['data']} "
            f"• Responsável: {resultados_atuais['responsavel'] or '-'}"
        )

    st.divider()

# ==========================================================
# FORMULÁRIO DE AVALIAÇÃO
# ==========================================================

with st.form(
    f"form_avaliacao_{arquivo_pet.stem}",
    clear_on_submit=False,
):
    st.subheader("📝 Notas por domínio")

    st.caption(
        "Use notas de 0 a 10. O peso define a importância de cada domínio no cálculo final."
    )

    valores_formulario = []

    for indice, item in enumerate(dominios):
        col_nome, col_nota, col_peso, col_resultado = st.columns(
            [3.5, 2, 2, 2]
        )

        with col_nome:
            st.markdown(
                f"**{indice + 1}. {item['dominio']}**"
            )

        with col_nota:
            nota = st.number_input(
                "Nota",
                min_value=0.0,
                max_value=10.0,
                value=float(item["nota"]),
                step=0.5,
                format="%.1f",
                key=f"nota_{arquivo_pet.stem}_{item['linha']}",
                label_visibility="collapsed",
            )

        with col_peso:
            peso = st.number_input(
                "Peso",
                min_value=0.1,
                max_value=10.0,
                value=float(item["peso"]),
                step=0.1,
                format="%.1f",
                key=f"peso_{arquivo_pet.stem}_{item['linha']}",
                label_visibility="collapsed",
            )

        with col_resultado:
            st.text_input(
                "Resultado",
                value=f"{nota * peso:.2f}",
                disabled=True,
                key=f"resultado_{arquivo_pet.stem}_{item['linha']}",
                label_visibility="collapsed",
            )

        valores_formulario.append(
            {
                "linha": item["linha"],
                "dominio": item["dominio"],
                "nota": float(nota),
                "peso": float(peso),
            }
        )

    st.divider()

    observacao = st.text_area(
        "Observação geral",
        value=observacao_atual,
        height=140,
        placeholder=(
            "Descreva informações clínicas, comportamentais ou observações importantes."
        ),
    )

    responsavel_padrao = (
        resultados_atuais["responsavel"]
        or pet["veterinario"]
    )

    responsavel = st.text_input(
        "Responsável pela avaliação *",
        value=responsavel_padrao,
        placeholder="Nome do veterinário ou profissional responsável",
    )

    salvar = st.form_submit_button(
        "💾 Salvar avaliação",
        type="primary",
        use_container_width=True,
    )

# ==========================================================
# PRÉVIA
# ==========================================================

previa = calcular_resultados(valores_formulario)

st.subheader("📈 Prévia do resultado")

col_p1, col_p2, col_p3, col_p4 = st.columns(4)

with col_p1:
    st.metric("Score geral", previa["score"])

with col_p2:
    st.metric("Média", previa["media"])

with col_p3:
    st.metric("Maior nota", previa["maior"])

with col_p4:
    st.metric("Classificação", previa["classificacao"])

st.progress(
    min(
        1.0,
        max(0.0, previa["media"] / 10),
    )
)

# ==========================================================
# SALVAMENTO
# ==========================================================

if salvar:
    if len(texto_limpo(responsavel)) < 2:
        st.error("❌ Informe o responsável pela avaliação.")
    else:
        try:
            resultados = salvar_avaliacao(
                arquivo=arquivo_pet,
                avaliacoes=valores_formulario,
                observacao=observacao,
                responsavel=responsavel,
            )

            st.success(
                f"✅ Avaliação de **{pet['nome']}** salva com sucesso!"
            )

            st.info(
                f"📊 Média: **{resultados['media']}** • "
                f"Classificação: **{resultados['classificacao']}**"
            )

            col_s1, col_s2 = st.columns(2)

            with col_s1:
                if st.button(
                    "📊 Abrir Dashboard",
                    use_container_width=True,
                ):
                    st.switch_page("pages/Dashboard.py")

            with col_s2:
                if st.button(
                    "🐾 Voltar para Pets",
                    use_container_width=True,
                ):
                    st.switch_page("pages/Pets.py")

        except Exception as erro:
            st.error("❌ Não foi possível salvar a avaliação.")
            st.exception(erro)

# ==========================================================
# RODAPÉ
# ==========================================================

st.divider()
st.caption(
    "DinoPetAnimal PRO • Avaliação veterinária individual em Excel"
)
